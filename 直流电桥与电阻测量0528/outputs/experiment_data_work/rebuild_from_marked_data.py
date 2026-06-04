from __future__ import annotations

import math
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(r"C:\Users\Administrator\Desktop\直流电桥与电阻测量0528")
INPUT = ROOT / "outputs" / "final" / "直流电桥与电阻测量_数据处理.xlsx"
OUTPUT = ROOT / "outputs" / "final" / "直流电桥与电阻测量_数据处理_带拟合参数表.xlsx"


def stdev(values: list[float]) -> float:
    mean = sum(values) / len(values)
    return math.sqrt(sum((x - mean) ** 2 for x in values) / (len(values) - 1))


def delta_b(value: float) -> float:
    # Decade-box class uncertainty from the PPT: 0.1% for >=10 Ω decades,
    # 0.5% for 1 Ω, 2% for 0.1 Ω, plus residual resistance 0.020 Ω.
    remaining = round(value, 1)
    total = 0.020
    for decade, rel in [
        (10000, 0.001),
        (1000, 0.001),
        (100, 0.001),
        (10, 0.001),
        (1, 0.005),
        (0.1, 0.02),
    ]:
        digit = int(math.floor((remaining + 1e-9) / decade))
        total += digit * decade * rel
        remaining -= digit * decade
    return total


def linreg(xs: list[float], ys: list[float]) -> tuple[float, float, float]:
    n = len(xs)
    mx = sum(xs) / n
    my = sum(ys) / n
    ssxx = sum((x - mx) ** 2 for x in xs)
    ssxy = sum((x - mx) * (y - my) for x, y in zip(xs, ys))
    slope = ssxy / ssxx
    intercept = my - slope * mx
    sst = sum((y - my) ** 2 for y in ys)
    ssr = sum((y - (intercept + slope * x)) ** 2 for x, y in zip(xs, ys))
    r2 = 1 - ssr / sst
    return slope, intercept, r2


def linreg_stats(xs: list[float], ys: list[float]) -> dict:
    n = len(xs)
    slope, intercept, r2 = linreg(xs, ys)
    mx = sum(xs) / n
    ssxx = sum((x - mx) ** 2 for x in xs)
    residuals = [y - (intercept + slope * x) for x, y in zip(xs, ys)]
    sse = sum(r * r for r in residuals)
    variance = sse / (n - 2)
    slope_se = math.sqrt(variance / ssxx)
    intercept_se = math.sqrt(variance * (1 / n + mx * mx / ssxx))
    pearson_r = math.copysign(math.sqrt(r2), slope)
    adjusted_r2 = 1 - (1 - r2) * (n - 1) / (n - 2)
    return {
        "slope": slope,
        "intercept": intercept,
        "r2": r2,
        "sse": sse,
        "slope_se": slope_se,
        "intercept_se": intercept_se,
        "pearson_r": pearson_r,
        "adjusted_r2": adjusted_r2,
    }


def bridge_calc(label: str, values: list[float], ratio: float, n_for_a: int | None = None) -> dict:
    n = len(values)
    n_factor = n_for_a or n
    mean = sum(values) / n
    sigma = stdev(values)
    if n_factor == 3:
        t_over_sqrt_n = 2.48
    elif n_factor == 6:
        t_over_sqrt_n = 2.571 / math.sqrt(6)
    else:
        t_over_sqrt_n = 2.48
    da = t_over_sqrt_n * sigma
    db_values = [delta_b(v) for v in values]
    db = max(db_values)
    u_r0 = math.hypot(da, db)
    rel_r0 = u_r0 / mean
    rx = mean * ratio
    rel_rx = math.hypot(0.0005, 0.0005, rel_r0)
    u_rx = rx * rel_rx
    return {
        "label": label,
        "n": n,
        "values": values,
        "mean": mean,
        "sigma": sigma,
        "da": da,
        "db_values": db_values,
        "db": db,
        "u_r0": u_r0,
        "rel_r0_pct": rel_r0 * 100,
        "ratio": ratio,
        "rx": rx,
        "rel_rx_pct": rel_rx * 100,
        "u_rx": u_rx,
    }


def final_bridge_text(item: dict) -> str:
    label = item["label"]
    rx = item["rx"]
    u = item["u_rx"]
    if label == "RA":
        return f"({rx / 1000:.2f} ± {u / 1000:.2f}) kΩ"
    if label.startswith("RB"):
        return f"({rx:.1f} ± {u:.1f}) Ω"
    return f"({rx:.2f} ± {u:.2f}) Ω"


def setup_sheet(ws, widths: list[float]) -> None:
    ws.sheet_view.showGridLines = False
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"


def style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")


def header(ws, row: int, cols: int) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for c in range(1, cols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


source_wb = load_workbook(INPUT, data_only=False)
src_bridge = source_wb["惠斯通电桥"]
src_four = source_wb["四端法原始数据"]

yellow = PatternFill("solid", fgColor="FFFF00")
blue = PatternFill("solid", fgColor="D9EAF7")

ra_r0 = [float(src_bridge[f"{col}2"].value) for col in "EFG"]
rb_direct_r0 = [float(src_bridge[f"{col}3"].value) for col in "EFG"]
rb_exchange_r0 = [float(src_bridge[f"{col}4"].value) for col in "EFG"]
rc_r0 = [float(src_bridge[f"{col}5"].value) for col in "EFG"]

bridge_results = [
    bridge_calc("RA", ra_r0, 100),
    bridge_calc("RB-交换前", rb_direct_r0, 1),
    bridge_calc("RB-交换后", rb_exchange_r0, 1),
    bridge_calc("RB-交换法(6次)", rb_direct_r0 + rb_exchange_r0, 1),
    bridge_calc("RC", rc_r0, 0.1),
]
summary_bridge = [bridge_results[0], bridge_results[3], bridge_results[4]]

samples: dict[str, dict] = {}
for row in range(2, 35):
    sample = src_four[f"A{row}"].value
    if sample is None:
        continue
    samples.setdefault(sample, {"I": [], "U": [], "d": float(src_four[f"E{row}"].value), "L": float(src_four[f"F{row}"].value)})
    samples[sample]["I"].append(float(src_four[f"C{row}"].value))
    samples[sample]["U"].append(float(src_four[f"D{row}"].value))

fit_results = {}
for sample, data in samples.items():
    stats = linreg_stats(data["I"], data["U"])
    slope = stats["slope"]
    intercept = stats["intercept"]
    r2 = stats["r2"]
    area = math.pi * (data["d"] * 1e-3) ** 2 / 4
    rho = slope * area / (data["L"] * 1e-3)
    fit_results[sample] = {
        "R": slope,
        "intercept": intercept,
        "r2": r2,
        "sse": stats["sse"],
        "slope_se": stats["slope_se"],
        "intercept_se": stats["intercept_se"],
        "pearson_r": stats["pearson_r"],
        "adjusted_r2": stats["adjusted_r2"],
        "area": area,
        "rho": rho,
    }

wb = Workbook()
ws_summary = wb.active
ws_summary.title = "结果汇总"
ws_bridge = wb.create_sheet("惠斯通电桥")
ws_raw = wb.create_sheet("四端法原始数据")
ws_fit = wb.create_sheet("四端法拟合图")
ws_note = wb.create_sheet("计算说明")

setup_sheet(ws_summary, [14, 18, 24, 16, 16, 18, 34])
setup_sheet(ws_bridge, [12, 14, 12, 12, 12, 12, 12, 10, 10, 10, 14, 10, 12, 12, 12, 12, 14, 14, 12, 24])
setup_sheet(ws_raw, [10, 8, 12, 12, 14, 14])
setup_sheet(ws_fit, [12, 14, 14, 12, 14, 16, 22, 2, 12, 12, 12, 12, 12, 12, 12])
setup_sheet(ws_note, [24, 95])

ws_note.append(["项目", "说明"])
ws_note.append(["标黄修改", "已重新读取 Excel 内标黄单元格：惠斯通 ΔU=0.14 mV、RB 交换后 912.1/912.1/912.2、S-2 的 59.88/70.06、S-3 的 60.08。"])
ws_note.append(["RB 最终处理", "RB 的交换法结果采用交换前后共 6 个 R0 读数综合处理；详细页列出交换前 3 次、交换后 3 次，以及最终交换法 6 次结果。"])
ws_note.append(["有效数字", "不确定度取 1 到 2 位；最终平均值保留到与不确定度相同的末位。表格数字设置了固定小数位，避免 Excel 把 80.00 等末尾 0 省略。"])
ws_note.append(["四端法图表", "按照 PPT 要求绘制 U-I 伏安特性曲线：散点为原始数据，同色直线为最小二乘拟合，横轴为电流 I/mA，纵轴为电压 U/mV。"])
header(ws_note, 1, 2)
style_range(ws_note, 1, ws_note.max_row, 1, 2)
for row in ws_note.iter_rows(min_row=2, max_row=ws_note.max_row, min_col=2, max_col=2):
    for cell in row:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

ws_summary.append(["实验项目", "对象", "最终结果", "相对不确定度", "拟合电阻 R/Ω", "电阻率 ρ/(Ω·m)", "备注"])
for item in summary_bridge:
    ws_summary.append([
        "惠斯通电桥",
        "RB（交换法6次）" if item["label"] == "RB-交换法(6次)" else item["label"],
        final_bridge_text(item),
        item["rel_rx_pct"] / 100,
        None,
        None,
        "交换前后共 6 次 R0；RA/RC 使用 3 次 R0" if item["label"] == "RB-交换法(6次)" else "",
    ])
ws_summary.append([None] * 7)
for sample in ["S-1", "S-2", "S-3"]:
    data = samples[sample]
    fit = fit_results[sample]
    ws_summary.append([
        "四端法",
        sample,
        None,
        None,
        fit["R"],
        fit["rho"],
        f"d={data['d']:.3f} mm, L={data['L']:.2f} mm",
    ])
header(ws_summary, 1, 7)
style_range(ws_summary, 1, ws_summary.max_row, 1, 7)
for row in range(2, 5):
    ws_summary[f"D{row}"].number_format = "0.00%"
for row in range(6, 9):
    ws_summary[f"E{row}"].number_format = "0.00000"
    ws_summary[f"F{row}"].number_format = "0.00E+00"

ws_bridge.append(["Rx", "万用表粗测", "R1 (Ω)", "R2 (Ω)", "R0-1 (Ω)", "R0-2 (Ω)", "R0-3 (Ω)", "I (mA)", "ΔR0 (Ω)", "ΔU (mV)"])
raw_bridge = [
    ["RA", "109.9 kΩ", 10000, 100, *ra_r0, 9.366, 0.1, float(src_bridge["J2"].value)],
    ["RB-交换前", "912.4 Ω", 1000, 1000, *rb_direct_r0, 5.451, 0.1, float(src_bridge["J3"].value)],
    ["RB-交换后", "", "", "", *rb_exchange_r0, "", "", ""],
    ["RC", "51.15 Ω", 100, 1000, *rc_r0, 17.79, 0.1, float(src_bridge["J5"].value)],
]
for row in raw_bridge:
    ws_bridge.append(row)
ws_bridge.append([])
ws_bridge.append(["Rx", "n", "R0平均 (Ω)", "σ (Ω)", "t0.95/√n", "ΔA (Ω)", "ΔB1 (Ω)", "ΔB2 (Ω)", "ΔB3 (Ω)", "ΔB4 (Ω)", "ΔB5 (Ω)", "ΔB6 (Ω)", "ΔBmax (Ω)", "u_R0 (Ω)", "u_rR0", "R1/R2", "Rx平均 (Ω)", "u_rRx", "u_Rx (Ω)", "最终表示"])
for item in bridge_results:
    factor = 2.48 if item["n"] == 3 else 2.571 / math.sqrt(6)
    db_values = item["db_values"] + [""] * (6 - len(item["db_values"]))
    ws_bridge.append([
        item["label"],
        item["n"],
        item["mean"],
        item["sigma"],
        factor,
        item["da"],
        *db_values,
        item["db"],
        item["u_r0"],
        item["rel_r0_pct"] / 100,
        item["ratio"],
        item["rx"],
        item["rel_rx_pct"] / 100,
        item["u_rx"],
        final_bridge_text(item),
    ])
ws_bridge.append([])
ws_bridge.append(["灵敏度", "ΔU/(ΔR0/R0)", "ΔU (mV)", "ΔR0/R0"])
for label, mean, du in [
    ("RA", bridge_results[0]["mean"], float(src_bridge["J2"].value)),
    ("RB", bridge_results[3]["mean"], float(src_bridge["J3"].value)),
    ("RC", bridge_results[4]["mean"], float(src_bridge["J5"].value)),
]:
    dr = 0.1
    ws_bridge.append([label, du / (dr / mean), du, dr / mean])

header(ws_bridge, 1, 10)
header(ws_bridge, 7, 20)
header(ws_bridge, 14, 4)
style_range(ws_bridge, 1, 5, 1, 10)
style_range(ws_bridge, 7, 12, 1, 20)
style_range(ws_bridge, 14, 17, 1, 4)
for row in [3, 4]:
    for col in range(1, 11):
        ws_bridge.cell(row, col).fill = yellow
for row in range(2, 6):
    for col in range(3, 8):
        ws_bridge.cell(row, col).number_format = "0.0"
    for col in range(8, 11):
        ws_bridge.cell(row, col).number_format = "0.000"
for row in range(8, 13):
    for col in range(3, 20):
        ws_bridge.cell(row, col).number_format = "0.000"
    ws_bridge.cell(row, 15).number_format = "0.00%"
    ws_bridge.cell(row, 18).number_format = "0.00%"
for row in range(15, 18):
    ws_bridge.cell(row, 2).number_format = "0.000"
    ws_bridge.cell(row, 3).number_format = "0.000"
    ws_bridge.cell(row, 4).number_format = "0.000000"

ws_raw.append(["样品", "点号", "I (mA)", "U (mV)", "直径 d (mm)", "长度 L (mm)"])
for sample in ["S-1", "S-2", "S-3"]:
    data = samples[sample]
    for idx, (i_value, u_value) in enumerate(zip(data["I"], data["U"])):
        ws_raw.append([sample, idx, i_value, u_value, data["d"], data["L"]])
header(ws_raw, 1, 6)
style_range(ws_raw, 1, ws_raw.max_row, 1, 6)
for row in range(2, ws_raw.max_row + 1):
    ws_raw[f"C{row}"].number_format = "0.00"
    ws_raw[f"D{row}"].number_format = "0.00"
    ws_raw[f"E{row}"].number_format = "0.000"
    ws_raw[f"F{row}"].number_format = "0.00"
for row in [19, 20, 30]:
    ws_raw[f"C{row}"].fill = yellow

ws_fit.append(["样品", "拟合电阻 R/Ω", "截距 b/mV", "R²", "横截面积 S/m²", "电阻率 ρ/(Ω·m)", "拟合方程"])
for sample in ["S-1", "S-2", "S-3"]:
    fit = fit_results[sample]
    ws_fit.append([
        sample,
        fit["R"],
        fit["intercept"],
        fit["r2"],
        fit["area"],
        fit["rho"],
        f"U = {fit['R']:.5f} I {'+' if fit['intercept'] >= 0 else '-'} {abs(fit['intercept']):.5f}",
    ])
ws_fit.append([])
ws_fit.append(["I (mA)", "S-1 实测", "S-1 拟合", "S-2 实测", "S-2 拟合", "S-3 实测", "S-3 拟合"])
for idx in range(11):
    i_value = samples["S-1"]["I"][idx]
    row = [f"{i_value:.2f}"]
    for sample in ["S-1", "S-2", "S-3"]:
        fit = fit_results[sample]
        row.append(samples[sample]["U"][idx])
        row.append(fit["intercept"] + fit["R"] * samples[sample]["I"][idx])
    ws_fit.append(row)
header(ws_fit, 1, 7)
header(ws_fit, 6, 7)
style_range(ws_fit, 1, 4, 1, 7)
style_range(ws_fit, 6, 17, 1, 7)
for row in range(2, 5):
    ws_fit[f"B{row}"].number_format = "0.00000"
    ws_fit[f"C{row}"].number_format = "0.00000"
    ws_fit[f"D{row}"].number_format = "0.000000"
    ws_fit[f"E{row}"].number_format = "0.00E+00"
    ws_fit[f"F{row}"].number_format = "0.00E+00"
for row in range(7, 18):
    for col in range(1, 8):
        ws_fit.cell(row, col).number_format = "0.00"

chart = LineChart()
chart.title = "伏安特性曲线求电阻"
chart.x_axis.title = "电流 I (mA)"
chart.y_axis.title = "电压 U (mV)"
chart.y_axis.scaling.min = 0
chart.y_axis.scaling.max = 11
chart.y_axis.majorUnit = 1
chart.legend.position = "r"
chart.height = 13
chart.width = 23

colors = {
    "S-1": "4F81BD",
    "S-2": "C0504D",
    "S-3": "9BBB59",
}
markers = {
    "S-1": "circle",
    "S-2": "square",
    "S-3": "triangle",
}
line_dashes = {
    "S-1": "solid",
    "S-2": "dash",
    "S-3": "dashDot",
}
chart_data = Reference(ws_fit, min_col=2, max_col=7, min_row=6, max_row=17)
chart_categories = Reference(ws_fit, min_col=1, min_row=7, max_row=17)
chart.add_data(chart_data, titles_from_data=True)
chart.set_categories(chart_categories)
for idx, sample in enumerate(["S-1", "S-2", "S-3"]):
    measured = chart.series[idx * 2]
    fitted = chart.series[idx * 2 + 1]
    measured.tx = SeriesLabel(v=f"{sample} 实测")
    fitted.tx = SeriesLabel(v=f"{sample} 拟合")
    measured.marker.symbol = markers[sample]
    measured.marker.size = 7
    measured.graphicalProperties.line.noFill = True
    measured.marker.graphicalProperties.solidFill = colors[sample]
    measured.marker.graphicalProperties.line.solidFill = colors[sample]
    fitted.marker.symbol = "none"
    fitted.graphicalProperties.line.solidFill = colors[sample]
    fitted.graphicalProperties.line.width = 22000
    if line_dashes[sample] != "solid":
        fitted.graphicalProperties.line.dashStyle = line_dashes[sample]
ws_fit.add_chart(chart, "I2")

legend_start = 20
legend_headers = {
    "S-1": 1,
    "S-2": 4,
    "S-3": 7,
}
legend_rows = [
    "方程",
    "绘图",
    "权重",
    "截距",
    "斜率",
    "残差平方和",
    "Pearson's r",
    "R²(COD)",
    "调整后R²",
]
for sample, start_col in legend_headers.items():
    fit = fit_results[sample]
    values = [
        "U = a + kI",
        sample,
        "不加权",
        f"{fit['intercept']:.5f} ± {fit['intercept_se']:.5f} mV",
        f"{fit['R']:.5f} ± {fit['slope_se']:.5f} Ω",
        f"{fit['sse']:.9f}",
        f"{fit['pearson_r']:.6f}",
        f"{fit['r2']:.6f}",
        f"{fit['adjusted_r2']:.6f}",
    ]
    for offset, label in enumerate(legend_rows):
        row = legend_start + offset
        ws_fit.cell(row, start_col, label)
        ws_fit.cell(row, start_col + 1, values[offset])
    merge_col = start_col + 1
    ws_fit.column_dimensions[get_column_letter(start_col)].width = 14
    ws_fit.column_dimensions[get_column_letter(merge_col)].width = 26
    for row in range(legend_start, legend_start + len(legend_rows)):
        for col in range(start_col, start_col + 2):
            cell = ws_fit.cell(row, col)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.border = Border(
                left=Side(style="thin", color="A6A6A6"),
                right=Side(style="thin", color="A6A6A6"),
                top=Side(style="thin", color="A6A6A6"),
                bottom=Side(style="thin", color="A6A6A6"),
            )
            cell.alignment = Alignment(horizontal="left", vertical="center")
    ws_fit.cell(legend_start, start_col).font = Font(name="等线", bold=True)
    ws_fit.cell(legend_start, start_col + 1).font = Font(name="等线", bold=True)

for ws in [ws_summary, ws_bridge, ws_raw, ws_fit, ws_note]:
    for row in ws.iter_rows():
        for cell in row:
            cell.font = Font(name="等线", size=11, bold=cell.font.bold)
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal or "center",
                vertical=cell.alignment.vertical or "center",
                wrap_text=cell.alignment.wrap_text,
            )
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

wb.save(OUTPUT)
print(OUTPUT)
