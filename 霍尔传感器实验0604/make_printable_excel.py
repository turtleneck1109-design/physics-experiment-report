from pathlib import Path
import math

import numpy as np
import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


ROOT = Path(__file__).resolve().parent
INPUT = ROOT / "霍尔传感器实验数据记录表_已填写.xlsx"
PLOT_DIR = ROOT / "outputs" / "python_plots"
OUT_DIR = ROOT / "outputs" / "hall_analysis_print"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT = OUT_DIR / "霍尔传感器实验数据处理与图表_打印版.xlsx"


def as_float(value):
    if value is None or value == "":
        return None
    return float(str(value).strip())


def row_values(ws, row, start_col, end_col):
    values = []
    for col in range(start_col, end_col + 1):
        value = as_float(ws.cell(row=row, column=col).value)
        if value is not None:
            values.append(value)
    return np.array(values, dtype=float)


def finite_solenoid_b(x_m, current_a, turns=3000, length_m=0.26, diameter_m=0.025):
    mu0 = 4 * math.pi * 1e-7
    x_m = np.array(x_m, dtype=float)
    term1 = (length_m + 2 * x_m) / (2 * np.sqrt(diameter_m**2 + (length_m + 2 * x_m) ** 2))
    term2 = (length_m - 2 * x_m) / (2 * np.sqrt(diameter_m**2 + (length_m - 2 * x_m) ** 2))
    return mu0 * turns / length_m * current_a * (term1 + term2)


def least_squares_line(x, y):
    x_mean = np.mean(x)
    y_mean = np.mean(y)
    slope = np.sum((x - x_mean) * (y - y_mean)) / np.sum((x - x_mean) ** 2)
    intercept = y_mean - slope * x_mean
    y_fit = slope * x + intercept
    ss_res = np.sum((y - y_fit) ** 2)
    ss_tot = np.sum((y - y_mean) ** 2)
    r2 = 1 - ss_res / ss_tot
    pearson_r = np.corrcoef(x, y)[0, 1]
    n = len(x)
    sxx = np.sum((x - x_mean) ** 2)
    sigma2 = ss_res / (n - 2)
    slope_se = np.sqrt(sigma2 / sxx)
    intercept_se = np.sqrt(sigma2 * (1 / n + x_mean**2 / sxx))
    adj_r2 = 1 - (1 - r2) * (n - 1) / (n - 2)
    return slope, intercept, y_fit, ss_res, pearson_r, r2, adj_r2, slope_se, intercept_se


source_wb = load_workbook(INPUT, data_only=True)
source_ws = source_wb["填写模板"]

us = row_values(source_ws, 5, 2, 16)
u0 = row_values(source_ws, 6, 2, 16)
u = row_values(source_ws, 7, 2, 16)
im = row_values(source_ws, 13, 2, 11)
u_prime2_mv = row_values(source_ws, 14, 2, 11)
x_raw_cm = np.concatenate([
    row_values(source_ws, 20, 2, 18),
    row_values(source_ws, 27, 2, 18),
    row_values(source_ws, 33, 2, 10),
])
u_prime3_mv = np.concatenate([
    row_values(source_ws, 21, 2, 18),
    row_values(source_ws, 28, 2, 18),
    row_values(source_ws, 34, 2, 10),
])

length_m = 0.26
diameter_m = 0.025
delta_x_cm = 2.3
b0_025_t = float(finite_solenoid_b([0], 0.25, length_m=length_m, diameter_m=diameter_m)[0])
b0_025_mt = b0_025_t * 1000
k_vs_us = (u - u0) / b0_025_t
k_over_us = k_vs_us / us
k_us_fit = least_squares_line(us, k_vs_us)
b2_mt = finite_solenoid_b(np.zeros_like(im), im, length_m=length_m, diameter_m=diameter_m) * 1000
fit = least_squares_line(b2_mt, u_prime2_mv)
k_fit, intercept_mv, u_fit_mv, sse, pearson_r, r2, adj_r2, slope_se, intercept_se = fit
k_us_slope, k_us_intercept, k_us_fit_y, k_us_sse, k_us_r, k_us_r2, k_us_adj_r2, k_us_slope_se, k_us_intercept_se = k_us_fit
relative_error = (k_fit - 31.25) / 31.25 * 100
x_center_cm = x_raw_cm - (delta_x_cm + length_m * 100 / 2)
b_exp_mt = u_prime3_mv / k_fit
b_theory_mt = finite_solenoid_b(x_center_cm / 100, 0.25, length_m=length_m, diameter_m=diameter_m) * 1000
rel_dev = (b_exp_mt - b_theory_mt) / b_theory_mt * 100

wb = Workbook()
ws_summary = wb.active
ws_summary.title = "打印版_结果汇总"
ws_data = wb.create_sheet("打印版_数据处理")
ws_fig = wb.create_sheet("打印版_图表")
ws_raw = wb.create_sheet("原始数据")

blue = "1F4E79"
light_blue = "D9EAF7"
light_gray = "F2F2F2"
white = "FFFFFF"
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)


def setup_print(ws, orientation="portrait"):
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = orientation
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = 0.35
    ws.page_margins.right = 0.35
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.2
    ws.page_margins.footer = 0.2
    ws.freeze_panes = "A2"


def title(ws, text, row, start_col, end_col):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.font = Font(name="Microsoft YaHei", size=16, bold=True, color=white)
    cell.fill = PatternFill("solid", fgColor=blue)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 28


def section(ws, text, row, start_col, end_col):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row, start_col, text)
    cell.font = Font(name="Microsoft YaHei", size=12, bold=True, color=blue)
    cell.fill = PatternFill("solid", fgColor=light_blue)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22


def write_table(ws, start_row, start_col, headers, rows, widths=None, number_format="0.0000"):
    for j, h in enumerate(headers, start_col):
        c = ws.cell(start_row, j, h)
        c.font = Font(name="Microsoft YaHei", bold=True, color=blue)
        c.fill = PatternFill("solid", fgColor=light_blue)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    for i, row in enumerate(rows, start_row + 1):
        for j, value in enumerate(row, start_col):
            c = ws.cell(i, j, value)
            c.font = Font(name="Microsoft YaHei", size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = border
            if isinstance(value, (int, float, np.floating)):
                c.number_format = number_format
    if widths:
        for offset, width in enumerate(widths):
            ws.column_dimensions[get_column_letter(start_col + offset)].width = width


def style_used(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.font = cell.font.copy(name="Microsoft YaHei")


setup_print(ws_summary, "portrait")
setup_print(ws_data, "landscape")
setup_print(ws_fig, "portrait")
setup_print(ws_raw, "landscape")

title(ws_summary, "霍尔传感器实验数据处理结果汇总", 1, 1, 6)
section(ws_summary, "一、关键公式与参数", 3, 1, 6)
write_table(ws_summary, 4, 1, ["项目", "结果", "单位/说明"], [
    ["中心磁场 B0(Im=0.25A)", b0_025_mt, "mT"],
    ["实验一 K-Us拟合方程", "K = a + bUs", ""],
    ["实验一截距 a", k_us_intercept, ""],
    ["实验一斜率 b", k_us_slope, "V/(T·V)"],
    ["实验一 R²(COD)", k_us_r2, ""],
    ["实验二最小二乘拟合方程", "U' = a + kB", "U': mV, B: mT"],
    ["截距 a", intercept_mv, "mV"],
    ["灵敏度 K = k", k_fit, "V/T"],
    ["参考灵敏度", 31.25, "V/T"],
    ["相对误差", relative_error, "%"],
    ["Pearson's r", pearson_r, ""],
    ["R²(COD)", r2, ""],
    ["调整后 R²", adj_r2, ""],
], widths=[24, 22, 26])

section(ws_summary, "二、可直接写入报告的结论", 16, 1, 6)
conclusion = [
    "1. 实验二中 U' 与 B 具有良好的线性关系，最小二乘拟合得到 K = %.2f V/T，R² = %.6f。" % (k_fit, r2),
    "2. 与 SS495A 参考灵敏度 31.25 V/T 相比，相对误差为 %.2f%%。" % relative_error,
    "3. 实验三得到的 B(x) 分布呈中部近似平台、两端快速下降的形状，与有限长螺线管理论曲线一致。",
    "4. 边缘区域磁场变化快，位置读数和探头对准误差会被放大，因此边缘相对偏差较大是合理的。",
]
for i, line in enumerate(conclusion, 17):
    ws_summary.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)
    c = ws_summary.cell(i, 1, line)
    c.font = Font(name="Microsoft YaHei", size=11)
    c.alignment = Alignment(wrap_text=True, vertical="center")
    ws_summary.row_dimensions[i].height = 26

section(ws_summary, "三、打印提示", 23, 1, 6)
ws_summary.merge_cells("A24:F25")
ws_summary["A24"] = "完整数据表见“打印版_数据处理”，四张图见“打印版_图表”。本工作簿已设置为 A4 打印、按页宽缩放。"
ws_summary["A24"].font = Font(name="Microsoft YaHei", size=11)
ws_summary["A24"].alignment = Alignment(wrap_text=True, vertical="center")

for col, width in enumerate([24, 18, 18, 18, 18, 18], 1):
    ws_summary.column_dimensions[get_column_letter(col)].width = width

title(ws_data, "霍尔传感器实验数据处理明细", 1, 1, 10)
section(ws_data, "实验一：改变 Us，计算 K 与 K/Us", 3, 1, 7)
write_table(ws_data, 4, 1, ["Us/V", "U0/V", "U/V", "U-U0/V", "B0/mT", "K/(V/T)", "K/Us"], [
    [us[i], u0[i], u[i], u[i] - u0[i], b0_025_mt, k_vs_us[i], k_over_us[i]]
    for i in range(len(us))
], widths=[10, 10, 10, 12, 12, 12, 12])
write_table(ws_data, 4, 9, ["拟合项目", "结果"], [
    ["方程", "K = a + bUs"],
    ["权重", "不加权"],
    ["截距 a", "%.4f ± %.4f" % (k_us_intercept, k_us_intercept_se)],
    ["斜率 b", "%.4f ± %.4f" % (k_us_slope, k_us_slope_se)],
    ["残差平方和", k_us_sse],
    ["Pearson's r", k_us_r],
    ["R²(COD)", k_us_r2],
    ["调整后R²", k_us_adj_r2],
], widths=[16, 24])

section(ws_data, "实验二：U' 与 B 的最小二乘拟合", 22, 1, 10)
write_table(ws_data, 23, 1, ["Im/A", "U'=U-U0/mV", "B/mT", "拟合U'/mV", "残差/mV"], [
    [im[i], u_prime2_mv[i], b2_mt[i], u_fit_mv[i], u_prime2_mv[i] - u_fit_mv[i]]
    for i in range(len(im))
], widths=[10, 14, 12, 14, 12])
write_table(ws_data, 23, 7, ["拟合项目", "结果"], [
    ["方程", "U' = a + kB"],
    ["权重", "不加权"],
    ["截距 a", "%.5f ± %.5f mV" % (intercept_mv, intercept_se)],
    ["斜率 k", "%.5f ± %.5f V/T" % (k_fit, slope_se)],
    ["残差平方和", sse],
    ["Pearson's r", pearson_r],
    ["R²(COD)", r2],
    ["调整后R²", adj_r2],
], widths=[16, 26])

section(ws_data, "实验三：B(x) 实验值与理论值", 37, 1, 10)
write_table(ws_data, 38, 1, ["原刻度X/cm", "中心坐标x/cm", "U'/mV", "B实验/mT", "B理论/mT", "相对偏差/%"], [
    [x_raw_cm[i], x_center_cm[i], u_prime3_mv[i], b_exp_mt[i], b_theory_mt[i], rel_dev[i]]
    for i in range(len(x_raw_cm))
], widths=[12, 14, 12, 12, 12, 12])
ws_data.row_breaks.append(Break(id=37))
ws_data.print_title_rows = "1:4"

title(ws_fig, "霍尔传感器实验图表（Python 绘图）", 1, 1, 8)
figure_files = [
    ("实验一_K随Us变化.png", "A3"),
    ("实验一_K除以Us随Us变化.png", "A24"),
    ("实验二_Uprime_B最小二乘拟合.png", "A45"),
    ("实验三_Bx实验值与理论值对比.png", "A66"),
]
for name, anchor in figure_files:
    path = PLOT_DIR / name
    if path.exists():
        img = Image(str(path))
        img.width = 620
        img.height = 400
        ws_fig.add_image(img, anchor)
for c in range(1, 9):
    ws_fig.column_dimensions[get_column_letter(c)].width = 13
for r in range(1, 88):
    ws_fig.row_dimensions[r].height = 18
for break_row in [23, 44, 65]:
    ws_fig.row_breaks.append(Break(id=break_row))

title(ws_raw, "原始数据记录（未修改）", 1, 1, 18)
for r in range(1, 36):
    for c in range(1, 19):
        target = ws_raw.cell(r + 2, c, source_ws.cell(r, c).value)
        target.font = Font(name="Microsoft YaHei", size=9)
        target.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        target.border = border
for c in range(1, 19):
    ws_raw.column_dimensions[get_column_letter(c)].width = 9
ws_raw.column_dimensions["A"].width = 18
ws_raw.print_title_rows = "1:3"

for sheet in wb.worksheets:
    style_used(sheet)

wb.save(OUTPUT)
print(OUTPUT)
