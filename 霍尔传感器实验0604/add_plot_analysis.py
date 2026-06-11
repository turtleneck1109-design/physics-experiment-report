from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parent
INPUT = ROOT / "outputs" / "hall_analysis_print" / "霍尔传感器实验数据处理与图表_打印版.xlsx"
OUTPUT = INPUT

wb = openpyxl.load_workbook(INPUT)
if "打印版_图表分析" in wb.sheetnames:
    del wb["打印版_图表分析"]
ws = wb.create_sheet("打印版_图表分析", 2)

blue = "1F4E79"
light_blue = "D9EAF7"
light_gray = "F2F2F2"
white = "FFFFFF"
thin = Side(style="thin", color="B7B7B7")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

ws.page_setup.paperSize = ws.PAPERSIZE_A4
ws.page_setup.orientation = "portrait"
ws.sheet_properties.pageSetUpPr.fitToPage = True
ws.page_setup.fitToWidth = 1
ws.page_setup.fitToHeight = 0
ws.page_margins.left = 0.35
ws.page_margins.right = 0.35
ws.page_margins.top = 0.45
ws.page_margins.bottom = 0.45

ws.merge_cells("A1:E1")
ws["A1"] = "图表分析与异常点说明"
ws["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color=white)
ws["A1"].fill = PatternFill("solid", fgColor=blue)
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 30

headers = ["图表", "主要现象", "结论", "异常/需说明点", "可能原因"]
for col, header in enumerate(headers, 1):
    cell = ws.cell(3, col, header)
    cell.font = Font(name="Microsoft YaHei", bold=True, color=blue)
    cell.fill = PatternFill("solid", fgColor=light_blue)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

rows = [
    [
        "实验一：K 随 Us 变化",
        "K 随工作电压 Us 增大整体上升；最小二乘拟合为 K = -5.3218 + 7.2166 Us，R²=0.9455。",
        "说明传感器输出差值 U-U0 随工作电压增大而增大，K 与 Us 近似线性相关，但线性程度受低电压异常点影响。",
        "Us=2.5 V 的 K 仅约 1.66 V/T，明显偏离后续趋势；Us=3.0 V、3.5 V 也偏低。",
        "SS495A 正常工作电压约为 4.5-10.5 V，低电压下传感器可能未进入正常线性工作状态；同时 Us=2.5 V 时 U0 数据接近 2.42 V，不符合零磁场输出约 Us/2 的通常特征，可能存在读数或接线状态异常。",
    ],
    [
        "实验一：K/Us 随 Us 变化",
        "K/Us 在 Us >= 3.5 V 后基本稳定在约 6.3-6.8，波动不大；低电压区上升很快。",
        "说明在正常供电范围内，灵敏度与工作电压近似成比例，K/Us 可作为判断传感器供电稳定性的辅助量。",
        "Us=2.5 V 和 3.0 V 的 K/Us 明显低于平台值，不宜作为线性区数据参与结论。",
        "低于额定工作电压时，传感器内部电路输出不稳定；零点输出和满量程响应均可能偏离线性比例关系。",
    ],
    [
        "实验二：U' 与 B 的最小二乘拟合",
        "实验点基本落在同一直线上；最小二乘拟合得到 K=30.0537 V/T，R²=0.999448，Pearson's r=0.999724。",
        "U'=U-U0 与 B 呈高度线性关系，实验测得的传感器灵敏度与参考值 31.25 V/T 相比，相对误差为 -3.83%，结果可信。",
        "Im=0 A 时 U'=-0.01 mV，接近零但略为负；拟合截距 a=3.4784 mV，并非严格为 0。",
        "差动测量仍存在零点残余、仪表零漂和调零误差；直流分压器或万用表读数也可能带来 mV 量级偏差。截距较小，不影响整体线性判断。",
    ],
    [
        "实验三：B(x) 实验值与理论值对比",
        "B(x) 曲线呈现中间近似平台、两端快速下降；中心附近实验值约 3.7 mT，理论值约 3.6 mT。",
        "实验曲线与有限长螺线管理论分布趋势一致，说明用实验二拟合得到的 K 换算磁场是合理的。",
        "两端区域相对误差较大；左端低场区域实验值相对理论值偏高，右端下降段也有一定偏差。",
        "边缘区域磁场梯度大，探头位置读数、螺线管起始刻度 2.3 cm 修正、霍尔探头未完全位于轴线、探头尺寸与摆放角度都会放大误差；中心平台区磁场变化缓慢，因此误差较小。",
    ],
]

for r, row in enumerate(rows, 4):
    for c, value in enumerate(row, 1):
        cell = ws.cell(r, c, value)
        cell.font = Font(name="Microsoft YaHei", size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border
        if c == 1:
            cell.fill = PatternFill("solid", fgColor=light_gray)
            cell.font = Font(name="Microsoft YaHei", size=10, bold=True, color=blue)
    ws.row_dimensions[r].height = 118

ws.merge_cells("A9:E9")
ws["A9"] = "可写入实验报告的综合结论"
ws["A9"].font = Font(name="Microsoft YaHei", size=12, bold=True, color=blue)
ws["A9"].fill = PatternFill("solid", fgColor=light_blue)
ws["A9"].alignment = Alignment(vertical="center")
ws["A9"].border = border

summary = (
    "实验结果表明，霍尔传感器输出电压差 U' 与磁场强度 B 具有良好的线性关系。"
    "由最小二乘法拟合得到灵敏度 K=30.05 V/T，与参考灵敏度 31.25 V/T 的相对误差为 -3.83%。"
    "螺线管内磁场分布实验曲线与理论曲线整体一致，中部磁场近似均匀，两端磁场迅速减小。"
    "低工作电压下的实验一数据以及螺线管边缘区域数据偏差较大，主要与传感器非额定供电、零点漂移、位置读数误差和边缘磁场梯度较大有关。"
)
ws.merge_cells("A10:E11")
ws["A10"] = summary
ws["A10"].font = Font(name="Microsoft YaHei", size=11)
ws["A10"].alignment = Alignment(vertical="top", wrap_text=True)
ws["A10"].border = border

widths = [20, 35, 35, 35, 40]
for i, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = width

wb.save(OUTPUT)
print(OUTPUT)
