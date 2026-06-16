from __future__ import annotations

import math
from pathlib import Path

import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from process_shm_data import DATA_FILE, G, FitResult, linear_fit, row_values


ROOT = Path(__file__).resolve().parent
CLEAN_FIG_DIR = ROOT / "output_figures_clean"
OUTPUT_XLSX = ROOT / "简谐振动实验数据处理结果_打印版.xlsx"
OUTPUT_XLSX_CLEAR = ROOT / "简谐振动实验数据处理结果_清晰打印版.xlsx"
OUTPUT_XLSX_FINAL = ROOT / "简谐振动实验数据处理结果_清晰重排版.xlsx"
OUTPUT_XLSX_BW = ROOT / "简谐振动实验数据处理结果_黑白紧凑打印版.xlsx"
OUTPUT_XLSX_BIG = ROOT / "简谐振动实验数据处理结果_大图简洁打印版.xlsx"


BLUE = "000000"
BLUE_FILL = "D9D9D9"
LIGHT_BLUE = "F2F2F2"
PALE = "FFFFFF"
GRID = "8A8A8A"
TEXT = "000000"
RED = "555555"
POINT_BLUE = "000000"

TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=16, color=TEXT)
SECTION_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color=TEXT)
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=10, color=TEXT)
BODY_FONT = Font(name="Microsoft YaHei", size=9, color=TEXT)
SMALL_FONT = Font(name="Microsoft YaHei", size=8, color=TEXT)
BLUE_FONT = Font(name="Microsoft YaHei", bold=True, size=8, color=TEXT)
RED_FONT = Font(name="Microsoft YaHei", bold=True, size=8, color=TEXT)
WHITE_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")

THIN = Side(style="thin", color=GRID)
MEDIUM_BLUE = Side(style="medium", color=BLUE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def setup_matplotlib() -> None:
    preferred = ["Microsoft YaHei", "SimHei", "Noto Sans CJK SC", "Arial Unicode MS"]
    available = {font.name for font in fm.fontManager.ttflist}
    for name in preferred:
        if name in available:
            plt.rcParams["font.sans-serif"] = [name]
            break
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["figure.facecolor"] = "white"


def set_page(ws, orientation: str = "landscape", fit_width: int = 1, fit_height: int = 1) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = orientation
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.35
    ws.page_margins.bottom = 0.35
    ws.page_margins.header = 0.12
    ws.page_margins.footer = 0.12


def style_block(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER


def merge_title(ws, cell_range: str, text: str, fill: str = BLUE) -> None:
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(":")[0]]
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = WHITE_FONT if fill == BLUE else SECTION_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(top=MEDIUM_BLUE, bottom=MEDIUM_BLUE)


def write_table(
    ws,
    start_row: int,
    start_col: int,
    headers: list[str],
    rows: list[list[object]],
    header_fill: str = BLUE_FILL,
) -> int:
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, header)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = BORDER
            if (r_idx - start_row) % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PALE)
    return start_row + len(rows)


def merged_table(
    ws,
    start_row: int,
    columns: list[tuple[str, int, int]],
    rows: list[list[object]],
    row_height: float = 28,
) -> int:
    for header, col1, col2 in columns:
        ws.merge_cells(start_row=start_row, start_column=col1, end_row=start_row, end_column=col2)
        cell = ws.cell(start_row, col1, header)
        cell.fill = PatternFill("solid", fgColor=BLUE_FILL)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for col in range(col1, col2 + 1):
            ws.cell(start_row, col).border = BORDER

    for r_offset, row_values_ in enumerate(rows, start=1):
        row_num = start_row + r_offset
        ws.row_dimensions[row_num].height = row_height
        for value, (_header, col1, col2) in zip(row_values_, columns):
            ws.merge_cells(start_row=row_num, start_column=col1, end_row=row_num, end_column=col2)
            cell = ws.cell(row_num, col1, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r_offset % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PALE)
            for col in range(col1, col2 + 1):
                c = ws.cell(row_num, col)
                c.border = BORDER
                if r_offset % 2 == 0:
                    c.fill = PatternFill("solid", fgColor=PALE)
    return start_row + len(rows)


def section(ws, row: int, text: str, last_col: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row, 1, text)
    cell.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    cell.font = SECTION_FONT
    cell.alignment = Alignment(vertical="center")
    cell.border = Border(top=MEDIUM_BLUE, bottom=THIN)
    ws.row_dimensions[row].height = 24


def load_data_and_results():
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["实验数据一页"]

    mass_g = row_values(ws, 4, 3, 7)
    force_n = mass_g / 1000.0 * G
    spring1_mass_g = float(ws.cell(row=5, column=3).value)
    spring2_mass_g = float(ws.cell(row=7, column=3).value)
    l1_m = row_values(ws, 6, 3, 7) / 1000.0
    l2_m = row_values(ws, 8, 3, 7) / 1000.0
    dl1_m = l1_m - l1_m[0]
    dl2_m = l2_m - l2_m[0]

    m_kg = row_values(ws, 13, 3, 7) / 1000.0
    t_s = row_values(ws, 14, 3, 7) / 1000.0

    amp_cm = row_values(ws, 18, 3, 6)
    amp_t_ms = row_values(ws, 19, 3, 6)

    amp_energy_cm = row_values(ws, 24, 3, 6)
    block_t_s = row_values(ws, 25, 3, 6) / 1000.0
    delta_x_m = 0.995 / 100.0
    vmax = delta_x_m / block_t_s

    inc1_m = row_values(ws, 29, 4, 7) / 1000.0
    inc1_t = row_values(ws, 30, 4, 7) / 1000.0
    inc2_m = row_values(ws, 31, 4, 7) / 1000.0
    inc2_t = row_values(ws, 32, 4, 7) / 1000.0

    fits: dict[str, FitResult] = {
        "弹簧1 F-ΔL": linear_fit(dl1_m, force_n),
        "弹簧2 F-ΔL": linear_fit(dl2_m, force_n),
        "水平 T²-M": linear_fit(m_kg, t_s**2),
        "斜面1 T²-M": linear_fit(inc1_m, inc1_t**2),
        "斜面2 T²-M": linear_fit(inc2_m, inc2_t**2),
        "A-T": linear_fit(amp_cm, amp_t_ms),
        "Vmax²-A²": linear_fit((amp_energy_cm / 100.0) ** 2, vmax**2),
    }

    data = {
        "static": (mass_g, force_n, l1_m, dl1_m, l2_m, dl2_m),
        "mass_period": (m_kg, t_s),
        "amplitude_period": (amp_cm, amp_t_ms),
        "energy": (amp_energy_cm, block_t_s, vmax),
        "inclines": (inc1_m, inc1_t, inc2_m, inc2_t),
        "spring_masses_g": (spring1_mass_g, spring2_mass_g),
    }
    return data, fits


def k_m0_from_period(fit: FitResult) -> tuple[float, float]:
    return 4.0 * math.pi**2 / fit.slope, fit.intercept / fit.slope


def build_summary_rows(data, fits):
    m1_g, m2_g = data["spring_masses_g"]
    k1 = fits["弹簧1 F-ΔL"].slope
    k2 = fits["弹簧2 F-ΔL"].slope
    m0_spring_g = (m1_g + m2_g) / 3.0
    k_tm, m0_tm = k_m0_from_period(fits["水平 T²-M"])
    k_inc1, m0_inc1 = k_m0_from_period(fits["斜面1 T²-M"])
    k_inc2, m0_inc2 = k_m0_from_period(fits["斜面2 T²-M"])
    amp_cm, amp_t_ms = data["amplitude_period"]
    amp_energy_cm, _block_t_s, _vmax = data["energy"]
    m_kg, _t_s = data["mass_period"]
    k_energy = fits["Vmax²-A²"].slope * (m_kg[0] + m0_tm)

    return [
        ["焦利秤法", f"k1={k1:.4f} N/m, k2={k2:.4f} N/m", f"k1+k2={k1 + k2:.4f} N/m；m0=(m1+m2)/3={m0_spring_g:.3f} g"],
        ["K值对比", f"焦利秤法 k1+k2={k1 + k2:.4f} N/m", f"周期法：水平 {k_tm:.4f}，斜面1 {k_inc1:.4f}，斜面2 {k_inc2:.4f} N/m"],
        ["水平 T²-M", f"K={k_tm:.4f} N/m", f"m0={m0_tm * 1000:.3f} g；T²={fits['水平 T²-M'].intercept:.6f}+{fits['水平 T²-M'].slope:.6f}M"],
        ["斜面1 T²-M", f"K={k_inc1:.4f} N/m", f"m0={m0_inc1 * 1000:.3f} g；h=1.240 cm，θ=0.822°"],
        ["斜面2 T²-M", f"K={k_inc2:.4f} N/m", f"m0={m0_inc2 * 1000:.3f} g；h=2.510 cm，θ=1.665°"],
        ["倾角影响", "周期与斜面倾角无明显关系", "倾角改变主要改变平衡位置；两斜面 K 值和同质量周期均很接近"],
        ["A-T 关系", f"T均值={np.mean(amp_t_ms):.3f} ms", f"s={np.std(amp_t_ms, ddof=1):.3f} ms；斜率={fits['A-T'].slope:.6f} ms/cm"],
        ["Vmax²-A²", f"斜率={fits['Vmax²-A²'].slope:.4f} s^-2", f"由 K=斜率(M+m0) 得 K={k_energy:.4f} N/m"],
    ]


def fit_rows(fits):
    units = {
        "弹簧1 F-ΔL": ("F=a+kΔL", "N", "N/m"),
        "弹簧2 F-ΔL": ("F=a+kΔL", "N", "N/m"),
        "水平 T²-M": ("T²=a+kM", "s²", "s²/kg"),
        "斜面1 T²-M": ("T²=a+kM", "s²", "s²/kg"),
        "斜面2 T²-M": ("T²=a+kM", "s²", "s²/kg"),
        "A-T": ("T=a+kA", "ms", "ms/cm"),
        "Vmax²-A²": ("Vmax²=a+kA²", "m²/s²", "s^-2"),
    }
    rows = []
    for name, fit in fits.items():
        equation, a_unit, k_unit = units[name]
        rows.append(
            [
                name,
                equation,
                f"{fit.intercept:.6g} ± {fit.intercept_se:.3g} {a_unit}",
                f"{fit.slope:.6g} ± {fit.slope_se:.3g} {k_unit}",
                f"{fit.ss_res:.6g}",
                f"{fit.pearson_r:.6f}",
                f"{fit.r2:.6f}",
                f"{fit.adj_r2:.6f}",
            ]
        )
    return rows


def make_clean_plot(
    x: np.ndarray,
    y: np.ndarray,
    fit: FitResult,
    xlabel: str,
    ylabel: str,
    filename: str,
    add_fit: bool = True,
    y_limits: tuple[float, float] | None = None,
    y_ticks: np.ndarray | None = None,
) -> Path:
    CLEAN_FIG_DIR.mkdir(exist_ok=True)
    fig, ax = plt.subplots(figsize=(5.0, 3.0), dpi=260)
    ax.scatter(x, y, s=36, color=f"#{POINT_BLUE}", edgecolors="white", linewidth=0.6, zorder=3)
    if add_fit:
        xx = np.linspace(float(np.min(x)), float(np.max(x)), 200)
        ax.plot(xx, fit.intercept + fit.slope * xx, color=f"#{RED}", linewidth=2.0, zorder=2)
    ax.set_xlabel(xlabel, fontsize=9)
    ax.set_ylabel(ylabel, fontsize=9)
    if y_limits is not None:
        ax.set_ylim(*y_limits)
    if y_ticks is not None:
        ax.set_yticks(y_ticks)
    ax.tick_params(labelsize=8)
    ax.grid(True, linestyle="--", linewidth=0.55, alpha=0.45)
    for spine in ax.spines.values():
        spine.set_color("#5B6470")
        spine.set_linewidth(0.8)
    fig.tight_layout(pad=0.6)
    path = CLEAN_FIG_DIR / filename
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    with PILImage.open(path) as image:
        grayscale = image.convert("L").convert("RGB")
        grayscale.save(path)
    return path


def generate_clean_figures(data, fits) -> list[dict[str, object]]:
    mass_g, force_n, _l1_m, dl1_m, _l2_m, dl2_m = data["static"]
    m_kg, t_s = data["mass_period"]
    amp_cm, amp_t_ms = data["amplitude_period"]
    amp_energy_cm, _block_t_s, vmax = data["energy"]
    inc1_m, inc1_t, inc2_m, inc2_t = data["inclines"]

    specs = [
        {
            "title": "图1 焦利秤法：弹簧1",
            "fit_key": "弹簧1 F-ΔL",
            "path": make_clean_plot(dl1_m, force_n, fits["弹簧1 F-ΔL"], "ΔL / m", "F / N", "plot_01_spring1.png"),
            "caption": f"F={fits['弹簧1 F-ΔL'].intercept:.4g}+{fits['弹簧1 F-ΔL'].slope:.4g}ΔL, R²={fits['弹簧1 F-ΔL'].r2:.6f}",
        },
        {
            "title": "图2 焦利秤法：弹簧2",
            "fit_key": "弹簧2 F-ΔL",
            "path": make_clean_plot(dl2_m, force_n, fits["弹簧2 F-ΔL"], "ΔL / m", "F / N", "plot_02_spring2.png"),
            "caption": f"F={fits['弹簧2 F-ΔL'].intercept:.4g}+{fits['弹簧2 F-ΔL'].slope:.4g}ΔL, R²={fits['弹簧2 F-ΔL'].r2:.6f}",
        },
        {
            "title": "图3 水平状态：T²-M",
            "fit_key": "水平 T²-M",
            "path": make_clean_plot(m_kg, t_s**2, fits["水平 T²-M"], "M / kg", "T² / s²", "plot_03_mass_period.png"),
            "caption": f"T²={fits['水平 T²-M'].intercept:.5f}+{fits['水平 T²-M'].slope:.5f}M, R²={fits['水平 T²-M'].r2:.6f}",
        },
        {
            "title": "图4 斜面1：T²-M",
            "fit_key": "斜面1 T²-M",
            "path": make_clean_plot(inc1_m, inc1_t**2, fits["斜面1 T²-M"], "M / kg", "T² / s²", "plot_04_incline1.png"),
            "caption": f"T²={fits['斜面1 T²-M'].intercept:.5f}+{fits['斜面1 T²-M'].slope:.5f}M, R²={fits['斜面1 T²-M'].r2:.6f}",
        },
        {
            "title": "图5 斜面2：T²-M",
            "fit_key": "斜面2 T²-M",
            "path": make_clean_plot(inc2_m, inc2_t**2, fits["斜面2 T²-M"], "M / kg", "T² / s²", "plot_05_incline2.png"),
            "caption": f"T²={fits['斜面2 T²-M'].intercept:.5f}+{fits['斜面2 T²-M'].slope:.5f}M, R²={fits['斜面2 T²-M'].r2:.6f}",
        },
        {
            "title": "图6 振幅 A 与周期 T",
            "fit_key": "A-T",
            "path": make_clean_plot(
                amp_cm,
                amp_t_ms,
                fits["A-T"],
                "A / cm",
                "T / ms",
                "plot_06_amplitude_period.png",
                y_limits=(1019.5, 1022.5),
                y_ticks=np.arange(1019.5, 1022.6, 0.5),
            ),
            "caption": f"T={fits['A-T'].intercept:.3f}{fits['A-T'].slope:+.5f}A, R²={fits['A-T'].r2:.6f}",
        },
        {
            "title": "图7 Vmax²-A²",
            "fit_key": "Vmax²-A²",
            "path": make_clean_plot((amp_energy_cm / 100.0) ** 2, vmax**2, fits["Vmax²-A²"], "A² / m²", "Vmax² / (m²/s²)", "plot_07_energy.png"),
            "caption": f"Vmax²={fits['Vmax²-A²'].intercept:.4f}+{fits['Vmax²-A²'].slope:.4f}A², R²={fits['Vmax²-A²'].r2:.6f}",
        },
    ]
    return specs


def build_overview_sheet(wb: Workbook, data, fits) -> None:
    ws = wb.active
    ws.title = "打印汇总"
    set_page(ws, "landscape", 1, 1)
    for col, width in enumerate([9, 9, 9, 13, 13, 13, 13, 14, 14, 14, 14, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    for row in range(1, 29):
        ws.row_dimensions[row].height = 20

    ws.merge_cells("A1:L1")
    ws["A1"] = "简谐振动实验数据处理结果"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:L2")
    ws["A2"] = "数据源：简谐振动实验数据整理_一页.xlsx    作图：Python    表格/图例：Excel 原生排版"
    ws["A2"].font = SMALL_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    section(ws, 4, "一、主要结果", 12)
    merged_table(
        ws,
        5,
        [("处理项目", 1, 3), ("结果", 4, 7), ("说明", 8, 12)],
        build_summary_rows(data, fits),
        row_height=25,
    )

    section(ws, 15, "二、计算公式", 12)
    formulas = [
        ["焦利秤法", "F = k ΔL"],
        ["质量-周期法", "T² = 4π²/(k1+k2) · M + 4π²/(k1+k2) · m0"],
        ["动能-势能法", "Vmax² = (k1+k2)/(M+m0) · A²，Vmax = Δx/t，Δx = 0.995 cm"],
    ]
    merged_table(
        ws,
        16,
        [("项目", 1, 3), ("公式", 4, 12)],
        formulas,
        row_height=24,
    )

    section(ws, 21, "三、结论摘要", 12)
    notes = [
        ["焦利秤法得到 k1+k2=4.7461 N/m；周期法得到水平 K=4.8624 N/m、斜面1 K=4.9087 N/m、斜面2 K=4.9035 N/m，两者整体接近。"],
        ["斜面倾角改变时，周期数据和拟合得到的 K 值变化很小，可认为周期与倾角无明显关系。"],
        ["A-T 拟合相关性很弱，说明在本实验范围内周期基本与振幅无关。"],
        ["Vmax² 与 A² 具有明显线性关系，符合简谐运动能量关系。"],
    ]
    merged_table(
        ws,
        22,
        [("说明", 1, 12)],
        notes,
        row_height=23,
    )

    ws.print_area = "A1:L27"


def build_fit_sheet(wb: Workbook, fits) -> None:
    ws = wb.create_sheet("拟合参数")
    set_page(ws, "landscape", 1, 1)
    ws.merge_cells("A1:H1")
    ws["A1"] = "线性拟合参数表（Excel 原生单元格）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30
    rows = fit_rows(fits)
    write_table(ws, 3, 1, ["拟合项目", "方程", "截距 a", "斜率 k", "残差平方和", "Pearson r", "R²", "调整后R²"], rows)
    widths = [18, 16, 25, 25, 15, 12, 12, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(3, 12):
        ws.row_dimensions[row].height = 28
    ws.freeze_panes = "A4"
    ws.print_area = "A1:H11"


def build_data_sheet(wb: Workbook, data) -> None:
    ws = wb.create_sheet("处理后数据")
    set_page(ws, "landscape", 1, 0)
    ws.merge_cells("A1:J1")
    ws["A1"] = "实验处理后数据（单位已换算）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 30

    r = 3
    mass_g, force_n, l1_m, dl1_m, l2_m, dl2_m = data["static"]
    section(ws, r, "一、焦利秤法", 10)
    rows = [[i + 1, mass_g[i], force_n[i], l1_m[i] * 1000, dl1_m[i], l2_m[i] * 1000, dl2_m[i]] for i in range(len(mass_g))]
    r = write_table(ws, r + 1, 1, ["序号", "砝码/g", "F/N", "L1/mm", "ΔL1/m", "L2/mm", "ΔL2/m"], rows) + 2

    m_kg, t_s = data["mass_period"]
    section(ws, r, "二、水平状态：M 与 T", 10)
    rows = [[i + 1, m_kg[i], t_s[i], t_s[i] ** 2] for i in range(len(m_kg))]
    r = write_table(ws, r + 1, 1, ["序号", "M/kg", "T/s", "T²/s²"], rows) + 2

    amp_cm, amp_t_ms = data["amplitude_period"]
    section(ws, r, "三、振幅 A 与周期 T", 10)
    rows = [[i + 1, amp_cm[i], amp_t_ms[i]] for i in range(len(amp_cm))]
    r = write_table(ws, r + 1, 1, ["序号", "A/cm", "T/ms"], rows) + 2

    amp_cm, block_t_s, vmax = data["energy"]
    section(ws, r, "四、振幅 A 与最大速度 Vmax", 10)
    rows = [[i + 1, amp_cm[i], (amp_cm[i] / 100.0) ** 2, block_t_s[i] * 1000, vmax[i], vmax[i] ** 2] for i in range(len(amp_cm))]
    r = write_table(ws, r + 1, 1, ["序号", "A/cm", "A²/m²", "挡光t/ms", "Vmax/(m/s)", "Vmax²/(m²/s²)"], rows) + 2

    inc1_m, inc1_t, inc2_m, inc2_t = data["inclines"]
    section(ws, r, "五、斜面状态：M 与 T", 10)
    rows = [[i + 1, inc1_m[i], inc1_t[i], inc1_t[i] ** 2, inc2_m[i], inc2_t[i], inc2_t[i] ** 2] for i in range(len(inc1_m))]
    r = write_table(ws, r + 1, 1, ["序号", "斜面1 M/kg", "斜面1 T/s", "斜面1 T²/s²", "斜面2 M/kg", "斜面2 T/s", "斜面2 T²/s²"], rows) + 1

    widths = [8, 14, 14, 14, 14, 16, 16, 12, 12, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    ws.freeze_panes = "A4"
    ws.print_area = f"A1:J{r}"


def legend_cells(ws, row: int, col: int, caption: str) -> None:
    blue_cell = ws.cell(row, col, "● 实验数据")
    blue_cell.font = BLUE_FONT
    red_cell = ws.cell(row, col + 2, "━ 线性拟合")
    red_cell.font = RED_FONT
    cap = ws.cell(row + 1, col, caption)
    cap.font = SMALL_FONT
    cap.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 6)
    for rr in range(row, row + 3):
        for cc in range(col, col + 7):
            ws.cell(rr, cc).border = BORDER
            if rr >= row + 1:
                ws.cell(rr, cc).fill = PatternFill("solid", fgColor=PALE)


def legend_side_panel(ws, row: int, col: int, caption: str) -> None:
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
    c = ws.cell(row, col, "图例")
    c.font = HEADER_FONT
    c.fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 3)
    c = ws.cell(row + 1, col, "● 实验数据")
    c.font = BLUE_FONT
    c.alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=row + 2, start_column=col, end_row=row + 2, end_column=col + 3)
    c = ws.cell(row + 2, col, "━ 线性拟合")
    c.font = RED_FONT
    c.alignment = Alignment(vertical="center")

    ws.merge_cells(start_row=row + 4, start_column=col, end_row=row + 8, end_column=col + 3)
    c = ws.cell(row + 4, col, caption)
    c.font = SMALL_FONT
    c.alignment = Alignment(vertical="top", wrap_text=True)

    for rr in range(row, row + 9):
        for cc in range(col, col + 4):
            cell = ws.cell(rr, cc)
            cell.border = BORDER
            if rr in (row, row + 4):
                cell.fill = PatternFill("solid", fgColor=PALE)


def add_image(ws, path: Path, anchor: str, width_px: int) -> None:
    image = XLImage(str(path))
    with PILImage.open(path) as pil_img:
        aspect = pil_img.height / pil_img.width
    image.width = width_px
    image.height = int(width_px * aspect)
    image.anchor = anchor
    ws.add_image(image)


def build_figures_sheet(wb: Workbook, figures, data, fits) -> None:
    ws = wb.create_sheet("图表总览")
    set_page(ws, "landscape", 1, 1)
    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 9.2
    for row in range(1, 76):
        ws.row_dimensions[row].height = 13.2
    ws.merge_cells("A1:P1")
    ws["A1"] = "Python 作图总览（大图简洁版）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    panels = [
        (3, 1), (3, 9),
        (21, 1), (21, 9),
        (39, 1), (39, 9),
        (57, 1),
    ]
    image_width = 365
    for fig, (start_row, start_col) in zip(figures, panels):
        ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 7)
        title_cell = ws.cell(start_row, start_col, fig["title"])
        title_cell.fill = PatternFill("solid", fgColor=BLUE_FILL)
        title_cell.font = HEADER_FONT
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        add_image(ws, fig["path"], f"{get_column_letter(start_col)}{start_row + 1}", image_width)
        caption_row = start_row + 15
        ws.merge_cells(start_row=caption_row, start_column=start_col, end_row=caption_row + 1, end_column=start_col + 7)
        caption_cell = ws.cell(caption_row, start_col, f"● 实验数据    ━ 线性拟合    {fig['caption']}")
        caption_cell.font = SMALL_FONT
        caption_cell.alignment = Alignment(vertical="top", wrap_text=True)
        caption_cell.fill = PatternFill("solid", fgColor="F7F7F7")

    # The eighth panel is a native-cell result card, not an image.
    card_row, card_col = 57, 9
    ws.merge_cells(start_row=card_row, start_column=card_col, end_row=card_row, end_column=card_col + 7)
    cell = ws.cell(card_row, card_col, "结果要点")
    cell.fill = PatternFill("solid", fgColor=BLUE_FILL)
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center")
    for cc in range(card_col, card_col + 8):
        ws.cell(card_row, cc).border = BORDER

    summary = build_summary_rows(data, fits)
    for i, row in enumerate(summary, start=card_row + 1):
        ws.cell(i, card_col, row[0])
        ws.cell(i, card_col + 1, row[1])
        ws.merge_cells(start_row=i, start_column=card_col + 1, end_row=i, end_column=card_col + 7)
        ws.row_dimensions[i].height = 19
        for cc in range(card_col, card_col + 8):
            c = ws.cell(i, cc)
            c.border = BORDER
            c.font = SMALL_FONT
            c.alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 0:
                c.fill = PatternFill("solid", fgColor=PALE)

    ws.print_area = "A1:P74"


def save_workbook() -> None:
    setup_matplotlib()
    data, fits = load_data_and_results()
    figures = generate_clean_figures(data, fits)

    wb = Workbook()
    build_overview_sheet(wb, data, fits)
    build_fit_sheet(wb, fits)
    build_data_sheet(wb, data)
    build_figures_sheet(wb, figures, data, fits)

    wb.save(OUTPUT_XLSX_BIG)
    print(OUTPUT_XLSX_BIG)


if __name__ == "__main__":
    save_workbook()
