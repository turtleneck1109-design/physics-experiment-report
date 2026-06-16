from __future__ import annotations

import math
from dataclasses import dataclass
from pathlib import Path

import matplotlib.font_manager as fm
import matplotlib.pyplot as plt
import numpy as np
import openpyxl


ROOT = Path(__file__).resolve().parent
DATA_FILE = ROOT / "简谐振动实验数据整理_一页.xlsx"
OUT_DIR = ROOT / "output_figures"
REPORT_FILE = ROOT / "data_processing_results.md"
G = 9.794


@dataclass
class FitResult:
    intercept: float
    slope: float
    intercept_se: float
    slope_se: float
    ss_res: float
    pearson_r: float
    r2: float
    adj_r2: float


def setup_fonts() -> None:
    preferred = [
        "Microsoft YaHei",
        "SimHei",
        "Noto Sans CJK SC",
        "Source Han Sans SC",
        "Arial Unicode MS",
    ]
    available = {font.name for font in fm.fontManager.ttflist}
    for name in preferred:
        if name in available:
            plt.rcParams["font.sans-serif"] = [name]
            break
    plt.rcParams["axes.unicode_minus"] = False
    plt.rcParams["mathtext.fontset"] = "dejavusans"


def row_values(ws, row: int, start_col: int, count: int) -> np.ndarray:
    values = [ws.cell(row=row, column=start_col + i).value for i in range(count)]
    return np.array([float(v) for v in values], dtype=float)


def linear_fit(x: np.ndarray, y: np.ndarray) -> FitResult:
    x = np.asarray(x, dtype=float)
    y = np.asarray(y, dtype=float)
    n = len(x)
    design = np.column_stack([np.ones(n), x])
    beta = np.linalg.lstsq(design, y, rcond=None)[0]
    y_hat = design @ beta
    residuals = y - y_hat
    ss_res = float(np.sum(residuals**2))
    ss_tot = float(np.sum((y - np.mean(y)) ** 2))
    r2 = 1.0 - ss_res / ss_tot if ss_tot else float("nan")
    adj_r2 = 1.0 - (1.0 - r2) * (n - 1) / (n - 2) if n > 2 else float("nan")
    pearson_r = float(np.corrcoef(x, y)[0, 1]) if n > 1 else float("nan")
    dof = max(n - 2, 1)
    variance = ss_res / dof
    covariance = variance * np.linalg.inv(design.T @ design)
    se = np.sqrt(np.diag(covariance))
    return FitResult(
        intercept=float(beta[0]),
        slope=float(beta[1]),
        intercept_se=float(se[0]),
        slope_se=float(se[1]),
        ss_res=ss_res,
        pearson_r=pearson_r,
        r2=float(r2),
        adj_r2=float(adj_r2),
    )


def fmt(value: float, digits: int = 6) -> str:
    if value == 0:
        return "0"
    if abs(value) < 1e-4 or abs(value) >= 1e5:
        return f"{value:.{digits}e}"
    return f"{value:.{digits}f}"


def fit_table_rows(
    fit: FitResult,
    equation: str,
    intercept_unit: str,
    slope_unit: str,
    x_label_short: str,
    y_label_short: str,
) -> list[list[str]]:
    return [
        ["方程", equation],
        ["绘图", f"{y_label_short} - {x_label_short}"],
        ["权重", "不加权"],
        ["截距 a", f"{fmt(fit.intercept)} ± {fmt(fit.intercept_se)} {intercept_unit}".strip()],
        ["斜率 k", f"{fmt(fit.slope)} ± {fmt(fit.slope_se)} {slope_unit}".strip()],
        ["残差平方和", fmt(fit.ss_res)],
        ["Pearson's r", fmt(fit.pearson_r)],
        ["R²(COD)", fmt(fit.r2)],
        ["调整后R²", fmt(fit.adj_r2)],
    ]


def add_fit_table(ax, rows: list[list[str]], title: str) -> None:
    ax.axis("off")
    ax.set_title(title, fontsize=12, fontweight="bold", pad=8)
    table = ax.table(
        cellText=rows,
        colLabels=["项目", "数值"],
        loc="center",
        cellLoc="left",
        colLoc="left",
        colWidths=[0.36, 0.64],
        bbox=[0.0, 0.0, 1.0, 0.94],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(8.4)
    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor("#B8C0CC")
        cell.set_linewidth(0.8)
        if row == 0:
            cell.set_facecolor("#EAF1FB")
            cell.set_text_props(weight="bold")
        elif row % 2 == 0:
            cell.set_facecolor("#F8FAFD")


def plot_fit(
    x: np.ndarray,
    y: np.ndarray,
    fit: FitResult,
    title: str,
    xlabel: str,
    ylabel: str,
    equation: str,
    intercept_unit: str,
    slope_unit: str,
    filename: str,
    x_short: str,
    y_short: str,
    annotation: str | None = None,
    y_limits: tuple[float, float] | None = None,
    y_ticks: np.ndarray | None = None,
) -> Path:
    fig = plt.figure(figsize=(11.6, 6.2), dpi=180)
    grid = fig.add_gridspec(1, 2, width_ratios=[1.7, 1.05], wspace=0.23)
    ax = fig.add_subplot(grid[0, 0])
    table_ax = fig.add_subplot(grid[0, 1])

    x_line = np.linspace(float(np.min(x)), float(np.max(x)), 200)
    y_line = fit.intercept + fit.slope * x_line
    ax.scatter(x, y, s=50, color="#2266AA", edgecolors="white", linewidth=0.9, label="实验数据", zorder=3)
    ax.plot(x_line, y_line, color="#C7433E", linewidth=2.2, label="线性拟合", zorder=2)
    ax.set_title(title, fontsize=14, fontweight="bold", pad=10)
    ax.set_xlabel(xlabel, fontsize=11)
    ax.set_ylabel(ylabel, fontsize=11)
    if y_limits is not None:
        ax.set_ylim(*y_limits)
    if y_ticks is not None:
        ax.set_yticks(y_ticks)
    ax.grid(True, linestyle="--", linewidth=0.7, alpha=0.45)
    ax.legend(frameon=True, loc="best")
    if annotation:
        ax.text(
            0.03,
            0.97,
            annotation,
            transform=ax.transAxes,
            va="top",
            ha="left",
            fontsize=9.5,
            bbox={"boxstyle": "round,pad=0.35", "facecolor": "white", "edgecolor": "#C8D0DC", "alpha": 0.95},
        )

    rows = fit_table_rows(fit, equation, intercept_unit, slope_unit, x_short, y_short)
    add_fit_table(table_ax, rows, "线性拟合结果")
    fig.tight_layout()
    path = OUT_DIR / filename
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    return path


def plot_amplitude_period(
    amplitude_cm: np.ndarray,
    period_ms: np.ndarray,
    fit: FitResult,
    mean_ms: float,
    std_ms: float,
) -> Path:
    annotation = f"平均周期 = {mean_ms:.3f} ms\n标准差 = {std_ms:.3f} ms\n相对标准差 = {std_ms / mean_ms * 100:.3f}%"
    return plot_fit(
        amplitude_cm,
        period_ms,
        fit,
        "振幅 A 与周期 T 的关系",
        "振幅 A / cm",
        "周期 T / ms",
        "T = a + kA",
        "ms",
        "ms/cm",
        "06_amplitude_period_fit.png",
        "A",
        "T",
        annotation,
        y_limits=(1019.5, 1022.5),
        y_ticks=np.arange(1019.5, 1022.6, 0.5),
    )


def plot_summary_table(summary_rows: list[list[str]]) -> Path:
    fig, ax = plt.subplots(figsize=(13.4, 6.6), dpi=180)
    ax.axis("off")
    ax.set_title("简谐振动实验数据处理结果汇总", fontsize=15, fontweight="bold", pad=12)
    table = ax.table(
        cellText=summary_rows,
        colLabels=["数据处理项目", "拟合/计算结果", "备注"],
        loc="center",
        cellLoc="left",
        colLoc="left",
        colWidths=[0.22, 0.42, 0.36],
        bbox=[0.0, 0.0, 1.0, 0.88],
    )
    table.auto_set_font_size(False)
    table.set_fontsize(9.2)
    for (row, col), cell in table.get_celld().items():
        cell.set_edgecolor("#B8C0CC")
        cell.set_linewidth(0.8)
        if row == 0:
            cell.set_facecolor("#EAF1FB")
            cell.set_text_props(weight="bold")
        elif row % 2 == 0:
            cell.set_facecolor("#F8FAFD")
    path = OUT_DIR / "00_results_summary.png"
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)
    return path


def markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    line = "| " + " | ".join(headers) + " |"
    sep = "| " + " | ".join(["---"] * len(headers)) + " |"
    body = ["| " + " | ".join(row) + " |" for row in rows]
    return "\n".join([line, sep, *body])


def main() -> None:
    setup_fonts()
    OUT_DIR.mkdir(exist_ok=True)

    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["实验数据一页"]

    mass_g = row_values(ws, 4, 3, 7)
    force_n = mass_g / 1000.0 * G
    spring1_mass_g = float(ws.cell(row=5, column=3).value)
    spring2_mass_g = float(ws.cell(row=7, column=3).value)
    l1_m = row_values(ws, 6, 3, 7) / 1000.0
    l2_m = row_values(ws, 8, 3, 7) / 1000.0
    dl1_m = l1_m - l1_m[0]
    dl2_m = l2_m - l2_m[0]

    oscillator_mass_kg = row_values(ws, 13, 3, 7) / 1000.0
    period_s = row_values(ws, 14, 3, 7) / 1000.0
    period2_s2 = period_s**2

    amplitude_cm = row_values(ws, 18, 3, 6)
    amplitude_period_ms = row_values(ws, 19, 3, 6)

    amplitude_energy_cm = row_values(ws, 24, 3, 6)
    block_time_s = row_values(ws, 25, 3, 6) / 1000.0
    delta_x_m = 0.995 / 100.0
    vmax_m_s = delta_x_m / block_time_s

    incline1_mass_kg = row_values(ws, 29, 4, 7) / 1000.0
    incline1_period_s = row_values(ws, 30, 4, 7) / 1000.0
    incline2_mass_kg = row_values(ws, 31, 4, 7) / 1000.0
    incline2_period_s = row_values(ws, 32, 4, 7) / 1000.0

    static1_fit = linear_fit(dl1_m, force_n)
    static2_fit = linear_fit(dl2_m, force_n)
    mass_period_fit = linear_fit(oscillator_mass_kg, period2_s2)
    incline1_fit = linear_fit(incline1_mass_kg, incline1_period_s**2)
    incline2_fit = linear_fit(incline2_mass_kg, incline2_period_s**2)
    amp_period_fit = linear_fit(amplitude_cm, amplitude_period_ms)
    energy_fit = linear_fit((amplitude_energy_cm / 100.0) ** 2, vmax_m_s**2)

    k1_static = static1_fit.slope
    k2_static = static2_fit.slope
    k_static_total = k1_static + k2_static
    m0_spring_kg = (spring1_mass_g + spring2_mass_g) / 3.0 / 1000.0

    def k_and_m0_from_period(fit: FitResult) -> tuple[float, float]:
        k_total = 4.0 * math.pi**2 / fit.slope
        m0_kg = fit.intercept / fit.slope
        return k_total, m0_kg

    k_tm, m0_tm_kg = k_and_m0_from_period(mass_period_fit)
    k_inc1, m0_inc1_kg = k_and_m0_from_period(incline1_fit)
    k_inc2, m0_inc2_kg = k_and_m0_from_period(incline2_fit)

    m_fixed_kg = float(oscillator_mass_kg[0])
    k_energy_from_tm_m0 = energy_fit.slope * (m_fixed_kg + m0_tm_kg)
    k_energy_from_spring_m0 = energy_fit.slope * (m_fixed_kg + m0_spring_kg)

    paths = []
    paths.append(
        plot_fit(
            dl1_m,
            force_n,
            static1_fit,
            "焦利秤法：弹簧1的劲度系数",
            "伸长量 ΔL / m",
            "拉力 F / N",
            "F = a + kΔL",
            "N",
            "N/m",
            "01_static_spring1_fit.png",
            "ΔL",
            "F",
            f"k1 = {k1_static:.4f} N/m",
        )
    )
    paths.append(
        plot_fit(
            dl2_m,
            force_n,
            static2_fit,
            "焦利秤法：弹簧2的劲度系数",
            "伸长量 ΔL / m",
            "拉力 F / N",
            "F = a + kΔL",
            "N",
            "N/m",
            "02_static_spring2_fit.png",
            "ΔL",
            "F",
            f"k2 = {k2_static:.4f} N/m",
        )
    )
    paths.append(
        plot_fit(
            oscillator_mass_kg,
            period2_s2,
            mass_period_fit,
            "振子质量 M 与周期平方 T² 的关系",
            "振子质量 M / kg",
            "周期平方 T² / s²",
            "T² = a + kM",
            "s²",
            "s²/kg",
            "03_mass_period_fit.png",
            "M",
            "T²",
            f"K = 4π²/k = {k_tm:.4f} N/m\nm0 = a/k = {m0_tm_kg * 1000:.3f} g",
        )
    )
    theta1_deg = math.degrees(math.asin(1.240 / 86.40))
    theta2_deg = math.degrees(math.asin(2.510 / 86.40))
    paths.append(
        plot_fit(
            incline1_mass_kg,
            incline1_period_s**2,
            incline1_fit,
            "斜面1：振子质量 M 与周期平方 T² 的关系",
            "振子质量 M / kg",
            "周期平方 T² / s²",
            "T² = a + kM",
            "s²",
            "s²/kg",
            "04_incline1_mass_period_fit.png",
            "M",
            "T²",
            f"h = 1.240 cm, θ = {theta1_deg:.3f}°\nK = {k_inc1:.4f} N/m, m0 = {m0_inc1_kg * 1000:.3f} g",
        )
    )
    paths.append(
        plot_fit(
            incline2_mass_kg,
            incline2_period_s**2,
            incline2_fit,
            "斜面2：振子质量 M 与周期平方 T² 的关系",
            "振子质量 M / kg",
            "周期平方 T² / s²",
            "T² = a + kM",
            "s²",
            "s²/kg",
            "05_incline2_mass_period_fit.png",
            "M",
            "T²",
            f"h = 2.510 cm, θ = {theta2_deg:.3f}°\nK = {k_inc2:.4f} N/m, m0 = {m0_inc2_kg * 1000:.3f} g",
        )
    )
    paths.append(
        plot_amplitude_period(
            amplitude_cm,
            amplitude_period_ms,
            amp_period_fit,
            float(np.mean(amplitude_period_ms)),
            float(np.std(amplitude_period_ms, ddof=1)),
        )
    )
    paths.append(
        plot_fit(
            (amplitude_energy_cm / 100.0) ** 2,
            vmax_m_s**2,
            energy_fit,
            "动能与势能关系：Vmax² 与 A² 的线性拟合",
            "振幅平方 A² / m²",
            "最大速度平方 Vmax² / (m²/s²)",
            "Vmax² = a + kA²",
            "m²/s²",
            "s^-2",
            "07_energy_vmax2_a2_fit.png",
            "A²",
            "Vmax²",
            f"Δx = 0.995 cm\nK = k(M+m0) = {k_energy_from_tm_m0:.4f} N/m",
        )
    )

    summary_rows = [
        [
            "焦利秤法",
            f"k1 = {k1_static:.4f} N/m; k2 = {k2_static:.4f} N/m; k1+k2 = {k_static_total:.4f} N/m",
            f"m0=(m1+m2)/3 = {m0_spring_kg * 1000:.3f} g",
        ],
        [
            "K值对比",
            f"焦利秤法: k1+k2 = {k_static_total:.4f} N/m",
            f"周期法: 水平 {k_tm:.4f}, 斜面1 {k_inc1:.4f}, 斜面2 {k_inc2:.4f} N/m",
        ],
        [
            "水平 T²-M",
            f"K = {k_tm:.4f} N/m; m0 = {m0_tm_kg * 1000:.3f} g",
            f"T² = {mass_period_fit.intercept:.6f} + {mass_period_fit.slope:.6f} M",
        ],
        [
            "斜面1 T²-M",
            f"K = {k_inc1:.4f} N/m; m0 = {m0_inc1_kg * 1000:.3f} g",
            f"h=1.240 cm, θ={theta1_deg:.3f}°",
        ],
        [
            "斜面2 T²-M",
            f"K = {k_inc2:.4f} N/m; m0 = {m0_inc2_kg * 1000:.3f} g",
            f"h=2.510 cm, θ={theta2_deg:.3f}°",
        ],
        [
            "倾角影响",
            "周期与斜面倾角无明显关系",
            "倾角主要改变平衡位置；两斜面 K 值和周期均接近",
        ],
        [
            "A-T 关系",
            f"T_mean = {np.mean(amplitude_period_ms):.3f} ms; s_T = {np.std(amplitude_period_ms, ddof=1):.3f} ms",
            f"线性斜率 = {amp_period_fit.slope:.6f} ms/cm\n周期基本与振幅无关",
        ],
        [
            "Vmax²-A²",
            f"拟合斜率 = {energy_fit.slope:.4f} s^-2; K = {k_energy_from_tm_m0:.4f} N/m",
            f"使用水平 T²-M 得到的 m0={m0_tm_kg * 1000:.3f} g\n若用(m1+m2)/3，则 K={k_energy_from_spring_m0:.4f} N/m",
        ],
    ]
    summary_path = plot_summary_table(summary_rows)
    paths.insert(0, summary_path)

    fit_rows = [
        [
            "弹簧1 F-ΔL",
            f"{static1_fit.intercept:.6g}",
            f"{static1_fit.intercept_se:.6g}",
            f"{static1_fit.slope:.6g}",
            f"{static1_fit.slope_se:.6g}",
            f"{static1_fit.pearson_r:.6f}",
            f"{static1_fit.r2:.6f}",
        ],
        [
            "弹簧2 F-ΔL",
            f"{static2_fit.intercept:.6g}",
            f"{static2_fit.intercept_se:.6g}",
            f"{static2_fit.slope:.6g}",
            f"{static2_fit.slope_se:.6g}",
            f"{static2_fit.pearson_r:.6f}",
            f"{static2_fit.r2:.6f}",
        ],
        [
            "水平 T²-M",
            f"{mass_period_fit.intercept:.6g}",
            f"{mass_period_fit.intercept_se:.6g}",
            f"{mass_period_fit.slope:.6g}",
            f"{mass_period_fit.slope_se:.6g}",
            f"{mass_period_fit.pearson_r:.6f}",
            f"{mass_period_fit.r2:.6f}",
        ],
        [
            "斜面1 T²-M",
            f"{incline1_fit.intercept:.6g}",
            f"{incline1_fit.intercept_se:.6g}",
            f"{incline1_fit.slope:.6g}",
            f"{incline1_fit.slope_se:.6g}",
            f"{incline1_fit.pearson_r:.6f}",
            f"{incline1_fit.r2:.6f}",
        ],
        [
            "斜面2 T²-M",
            f"{incline2_fit.intercept:.6g}",
            f"{incline2_fit.intercept_se:.6g}",
            f"{incline2_fit.slope:.6g}",
            f"{incline2_fit.slope_se:.6g}",
            f"{incline2_fit.pearson_r:.6f}",
            f"{incline2_fit.r2:.6f}",
        ],
        [
            "A-T",
            f"{amp_period_fit.intercept:.6g}",
            f"{amp_period_fit.intercept_se:.6g}",
            f"{amp_period_fit.slope:.6g}",
            f"{amp_period_fit.slope_se:.6g}",
            f"{amp_period_fit.pearson_r:.6f}",
            f"{amp_period_fit.r2:.6f}",
        ],
        [
            "Vmax²-A²",
            f"{energy_fit.intercept:.6g}",
            f"{energy_fit.intercept_se:.6g}",
            f"{energy_fit.slope:.6g}",
            f"{energy_fit.slope_se:.6g}",
            f"{energy_fit.pearson_r:.6f}",
            f"{energy_fit.r2:.6f}",
        ],
    ]

    report = [
        "# 简谐振动实验数据处理结果",
        "",
        "数据来自 `简谐振动实验数据整理_一页.xlsx`。所有作图和线性拟合均由 `process_shm_data.py` 完成，拟合均为不加权最小二乘。",
        "",
        "## 主要公式",
        "",
        "- 焦利秤法：`F = k ΔL`。",
        "- 质量-周期法：`T² = 4π²/(k1+k2) · M + 4π²/(k1+k2) · m0`。",
        "- 动能-势能法：`Vmax² = (k1+k2)/(M+m0) · A²`，其中 `Vmax = Δx/t`，`Δx = 0.995 cm`。",
        "",
        "## 物理结果汇总",
        "",
        markdown_table(["项目", "结果", "说明"], summary_rows),
        "",
        "## 线性拟合参数汇总",
        "",
        markdown_table(["拟合", "截距a", "a标准误", "斜率k", "k标准误", "Pearson r", "R²"], fit_rows),
        "",
        "## 输出图表",
        "",
        *[f"- `{path.relative_to(ROOT)}`" for path in paths],
        "",
    ]
    REPORT_FILE.write_text("\n".join(report), encoding="utf-8")

    print("Generated figures:")
    for path in paths:
        print(path)
    print(f"Report: {REPORT_FILE}")


if __name__ == "__main__":
    main()
