from __future__ import annotations

from pathlib import Path

import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from build_reference_style_excel import (
    BLACK,
    BODY_FONT,
    BORDER,
    FONT_NAME,
    GRID,
    HEADER_FILL,
    HEADER_FONT,
    LIGHT_FILL,
    MEDIUM,
    SMALL_FONT,
    TITLE_FONT,
    WHITE,
    figure_specs,
    k_m0_from_period,
    load_data_and_results,
)


ROOT = Path(__file__).resolve().parent
OUTPUT_DIR = ROOT / "outputs" / "reference_style_excel"
OUTPUT_XLSX = OUTPUT_DIR / "简谐振动实验图表图例合并一页_修正版.xlsx"


THIN = Side(style="thin", color=GRID)
OUTLINE = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def set_one_page(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.15
    ws.page_margins.right = 0.15
    ws.page_margins.top = 0.2
    ws.page_margins.bottom = 0.2
    ws.page_margins.header = 0.05
    ws.page_margins.footer = 0.05


def add_image(ws, image_path: Path, anchor: str, width_px: int) -> None:
    image = XLImage(str(image_path))
    with PILImage.open(image_path) as pil:
        aspect = pil.height / pil.width
    image.width = width_px
    image.height = int(width_px * aspect)
    image.anchor = anchor
    ws.add_image(image)


def compact_rows(rows: list[list[str]]) -> list[list[str]]:
    keep = {"方程", "绘图", "图例", "截距 a", "斜率 k", "Pearson's r", "R²(COD)", "物理量", "换算结果", "统计量", "K值", "结论"}
    title = rows[0][0]
    selected = [[label, value] for label, value in rows[1:] if label in keep]
    if len(selected) > 8:
        selected = selected[:8]
    return [[title, ""], *selected]


def write_small_table(ws, start_row: int, start_col: int, rows: list[list[str]], max_rows: int = 9) -> None:
    rows = rows[:max_rows]
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
    header = ws.cell(start_row, start_col, rows[0][0])
    header.fill = PatternFill("solid", fgColor=HEADER_FILL)
    header.font = HEADER_FONT
    header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for offset, row in enumerate(rows[1:], start=1):
        rr = start_row + offset
        label = ws.cell(rr, start_col, row[0])
        label.fill = PatternFill("solid", fgColor=LIGHT_FILL)
        label.font = SMALL_FONT
        label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(start_row=rr, start_column=start_col + 1, end_row=rr, end_column=start_col + 3)
        value = ws.cell(rr, start_col + 1, row[1])
        value.font = SMALL_FONT
        value.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rr in range(start_row, start_row + max_rows):
        ws.row_dimensions[rr].height = 13.5
        for cc in range(start_col, start_col + 4):
            cell = ws.cell(rr, cc)
            cell.border = BORDER
            if rr == start_row:
                cell.border = OUTLINE


def write_block(ws, start_row: int, start_col: int, spec: dict[str, object]) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 11)
    title = ws.cell(start_row, start_col, spec["title"])
    title.font = Font(name=FONT_NAME, bold=True, size=9, color=BLACK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor=WHITE)
    for cc in range(start_col, start_col + 12):
        ws.cell(start_row, cc).border = Border()

    write_small_table(ws, start_row + 1, start_col, compact_rows(spec["rows"]))
    add_image(ws, spec["image"], f"{get_column_letter(start_col + 5)}{start_row + 1}", 280)


def write_summary_block(ws, start_row: int, start_col: int, data, fits) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 11)
    cell = ws.cell(start_row, start_col, "主要结果汇总")
    cell.font = Font(name=FONT_NAME, bold=True, size=9, color=BLACK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    for cc in range(start_col, start_col + 12):
        ws.cell(start_row, cc).border = Border()

    k_tm, m0_tm = k_m0_from_period(fits["水平 T²-M"])
    k_inc1, m0_inc1 = k_m0_from_period(fits["斜面1 T²-M"])
    k_inc2, m0_inc2 = k_m0_from_period(fits["斜面2 T²-M"])
    rows = [
        ["项目", "结果"],
        ["焦利秤法", f"k1+k2={fits['弹簧1 F-ΔL'].slope + fits['弹簧2 F-ΔL'].slope:.4f} N/m"],
        ["水平周期法", f"K={k_tm:.4f} N/m, m0={m0_tm * 1000:.3f} g"],
        ["斜面1", f"K={k_inc1:.4f} N/m, m0={m0_inc1 * 1000:.3f} g"],
        ["斜面2", f"K={k_inc2:.4f} N/m, m0={m0_inc2 * 1000:.3f} g"],
        ["振幅-周期", f"Tmean={np.mean(data['amp_t_ms']):.3f} ms"],
        ["结论", "倾角影响不明显；周期基本与振幅无关"],
    ]
    for r_offset, row in enumerate(rows, start=1):
        rr = start_row + r_offset
        ws.cell(rr, start_col, row[0])
        ws.merge_cells(start_row=rr, start_column=start_col + 1, end_row=rr, end_column=start_col + 11)
        ws.cell(rr, start_col + 1, row[1])
        for cc in range(start_col, start_col + 12):
            c = ws.cell(rr, cc)
            c.font = SMALL_FONT if r_offset else BODY_FONT
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER
            if r_offset == 1 or cc == start_col:
                c.fill = PatternFill("solid", fgColor=LIGHT_FILL)


def build_workbook() -> None:
    data, fits = load_data_and_results()
    specs = figure_specs(data, fits)

    wb = Workbook()
    ws = wb.active
    ws.title = "合并一页"
    set_one_page(ws)

    widths = [
        7.4, 7.4, 7.4, 7.4, 5.8, 5.8, 5.8, 5.8, 5.8, 5.8, 5.8, 5.8,
        1.0,
        7.4, 7.4, 7.4, 7.4, 5.8, 5.8, 5.8, 5.8, 5.8, 5.8, 5.8, 5.8,
    ]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for rr in range(1, 58):
        ws.row_dimensions[rr].height = 13.5

    ws.merge_cells("A1:Y1")
    ws["A1"] = "简谐振动实验图表与详细图例（合并一页）"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    positions = [
        (3, 1), (3, 14),
        (16, 1), (16, 14),
        (29, 1), (29, 14),
        (42, 1),
    ]
    for spec, (row, col) in zip(specs, positions):
        write_block(ws, row, col, spec)
    write_summary_block(ws, 42, 14, data, fits)

    ws.print_area = "A1:Y54"
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    build_workbook()
