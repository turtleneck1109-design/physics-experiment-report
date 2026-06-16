from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from process_shm_data import DATA_FILE, G, FitResult, linear_fit, row_values


ROOT = Path(__file__).resolve().parent
CLEAN_FIG_DIR = ROOT / "output_figures_clean"
OUTPUT_DIR = ROOT / "outputs" / "reference_style_excel"
OUTPUT_XLSX = OUTPUT_DIR / "简谐振动实验图表图例参考版.xlsx"

FONT_NAME = "Microsoft YaHei"
GRID = "808080"
BLACK = "000000"
HEADER_FILL = "D9D9D9"
LIGHT_FILL = "EFEFEF"
WHITE = "FFFFFF"

TITLE_FONT = Font(name=FONT_NAME, bold=True, size=13, color=BLACK)
HEADER_FONT = Font(name=FONT_NAME, bold=True, size=9, color=BLACK)
BODY_FONT = Font(name=FONT_NAME, size=8, color=BLACK)
SMALL_FONT = Font(name=FONT_NAME, size=7, color=BLACK)

THIN = Side(style="thin", color=GRID)
MEDIUM = Side(style="medium", color=BLACK)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
OUTLINE = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)


def fmt(value: float, digits: int = 6) -> str:
    if value == 0:
        return "0"
    if abs(value) < 1e-4 or abs(value) >= 1e5:
        return f"{value:.{digits}e}"
    return f"{value:.{digits}f}".rstrip("0").rstrip(".")


def fit_row_table(
    name: str,
    fit: FitResult,
    equation: str,
    plot_name: str,
    intercept_unit: str,
    slope_unit: str,
    extra_rows: list[list[str]] | None = None,
) -> list[list[str]]:
    rows = [
        ["方程", equation],
        ["绘图", plot_name],
        ["图例", "● 实验数据；━ 线性拟合"],
        ["权重", "不加权"],
        ["截距 a", f"{fmt(fit.intercept)} ± {fmt(fit.intercept_se)} {intercept_unit}".strip()],
        ["斜率 k", f"{fmt(fit.slope)} ± {fmt(fit.slope_se)} {slope_unit}".strip()],
        ["残差平方和", fmt(fit.ss_res)],
        ["Pearson's r", fmt(fit.pearson_r)],
        ["R²(COD)", fmt(fit.r2)],
        ["调整后R²", fmt(fit.adj_r2)],
    ]
    if extra_rows:
        rows.extend(extra_rows)
    return [[name, ""], *rows]


def load_data_and_results():
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["实验数据一页"]

    mass_g = row_values(ws, 4, 3, 7)
    force_n = mass_g / 1000.0 * G
    spring1_mass_g = float(ws.cell(row=5, column=3).value)
    spring2_mass_g = float(ws.cell(row=7, column=3).value)
    l1_m = row_values(ws, 6, 3, 7) / 1000.0
    l2_m = row_values(ws, 8, 3, 7) / 1000.0
    dl1_m = l1_m - l1_m[0]
    dl2_m = l2_m - l2_m[0]

    m_kg = row_values(ws, 13, 3, 7) / 1000.0
    t_s = row_values(ws, 14, 3, 7) / 1000.0

    amp_cm = row_values(ws, 18, 3, 6)
    amp_t_ms = row_values(ws, 19, 3, 6)

    amp_energy_cm = row_values(ws, 24, 3, 6)
    block_t_s = row_values(ws, 25, 3, 6) / 1000.0
    delta_x_m = 0.995 / 100.0
    vmax = delta_x_m / block_t_s

    inc1_m = row_values(ws, 29, 4, 7) / 1000.0
    inc1_t = row_values(ws, 30, 4, 7) / 1000.0
    inc2_m = row_values(ws, 31, 4, 7) / 1000.0
    inc2_t = row_values(ws, 32, 4, 7) / 1000.0

    fits = {
        "弹簧1 F-ΔL": linear_fit(dl1_m, force_n),
        "弹簧2 F-ΔL": linear_fit(dl2_m, force_n),
        "水平 T²-M": linear_fit(m_kg, t_s**2),
        "斜面1 T²-M": linear_fit(inc1_m, inc1_t**2),
        "斜面2 T²-M": linear_fit(inc2_m, inc2_t**2),
        "A-T": linear_fit(amp_cm, amp_t_ms),
        "Vmax²-A²": linear_fit((amp_energy_cm / 100.0) ** 2, vmax**2),
    }
    data = {
        "spring_masses_g": (spring1_mass_g, spring2_mass_g),
        "m_kg": m_kg,
        "amp_t_ms": amp_t_ms,
    }
    return data, fits


def k_m0_from_period(fit: FitResult) -> tuple[float, float]:
    return 4.0 * math.pi**2 / fit.slope, fit.intercept / fit.slope


def figure_specs(data, fits) -> list[dict[str, object]]:
    m1_g, m2_g = data["spring_masses_g"]
    m0_spring_kg = (m1_g + m2_g) / 3.0 / 1000.0
    k_tm, m0_tm = k_m0_from_period(fits["水平 T²-M"])
    k_inc1, m0_inc1 = k_m0_from_period(fits["斜面1 T²-M"])
    k_inc2, m0_inc2 = k_m0_from_period(fits["斜面2 T²-M"])
    mean_t = float(np.mean(data["amp_t_ms"]))
    std_t = float(np.std(data["amp_t_ms"], ddof=1))
    k_energy = fits["Vmax²-A²"].slope * (float(data["m_kg"][0]) + m0_tm)
    k_energy_alt = fits["Vmax²-A²"].slope * (float(data["m_kg"][0]) + m0_spring_kg)

    return [
        {
            "title": "图1 焦利秤法：弹簧1劲度系数",
            "image": CLEAN_FIG_DIR / "plot_01_spring1.png",
            "rows": fit_row_table(
                "F = a + kΔL",
                fits["弹簧1 F-ΔL"],
                "F = a + kΔL",
                "F - ΔL",
                "N",
                "N/m",
                [["物理量", "k1 = 2.2206 N/m"], ["结论", "满足胡克定律"]],
            ),
        },
        {
            "title": "图2 焦利秤法：弹簧2劲度系数",
            "image": CLEAN_FIG_DIR / "plot_02_spring2.png",
            "rows": fit_row_table(
                "F = a + kΔL",
                fits["弹簧2 F-ΔL"],
                "F = a + kΔL",
                "F - ΔL",
                "N",
                "N/m",
                [["物理量", "k2 = 2.5255 N/m"], ["结论", "弹簧2较硬"]],
            ),
        },
        {
            "title": "图3 水平状态：T²-M 关系",
            "image": CLEAN_FIG_DIR / "plot_03_mass_period.png",
            "rows": fit_row_table(
                "T² = a + kM",
                fits["水平 T²-M"],
                "T² = a + kM",
                "T² - M",
                "s²",
                "s²/kg",
                [[
                    "换算结果",
                    f"K=4π²/k={k_tm:.4f} N/m",
                ], ["附加质量", f"m0=a/k={m0_tm * 1000:.3f} g"]],
            ),
        },
        {
            "title": "图4 斜面1：T²-M 关系",
            "image": CLEAN_FIG_DIR / "plot_04_incline1.png",
            "rows": fit_row_table(
                "T² = a + kM",
                fits["斜面1 T²-M"],
                "T² = a + kM",
                "T² - M",
                "s²",
                "s²/kg",
                [["斜面", "h=1.240 cm, θ=0.822°"], ["换算结果", f"K={k_inc1:.4f} N/m, m0={m0_inc1 * 1000:.3f} g"]],
            ),
        },
        {
            "title": "图5 斜面2：T²-M 关系",
            "image": CLEAN_FIG_DIR / "plot_05_incline2.png",
            "rows": fit_row_table(
                "T² = a + kM",
                fits["斜面2 T²-M"],
                "T² = a + kM",
                "T² - M",
                "s²",
                "s²/kg",
                [["斜面", "h=2.510 cm, θ=1.665°"], ["换算结果", f"K={k_inc2:.4f} N/m, m0={m0_inc2 * 1000:.3f} g"]],
            ),
        },
        {
            "title": "图6 振幅 A 与周期 T 的关系",
            "image": CLEAN_FIG_DIR / "plot_06_amplitude_period.png",
            "rows": fit_row_table(
                "T = a + kA",
                fits["A-T"],
                "T = a + kA",
                "T - A",
                "ms",
                "ms/cm",
                [["统计量", f"Tmean={mean_t:.3f} ms, s={std_t:.3f} ms"], ["结论", "周期基本与振幅无关"]],
            ),
        },
        {
            "title": "图7 Vmax² 与 A² 的能量关系",
            "image": CLEAN_FIG_DIR / "plot_07_energy.png",
            "rows": fit_row_table(
                "Vmax² = a + kA²",
                fits["Vmax²-A²"],
                "Vmax² = a + kA²",
                "Vmax² - A²",
                "m²/s²",
                "s^-2",
                [["计算", "Vmax=Δx/t, Δx=0.995 cm"], ["K值", f"K={k_energy:.4f} N/m；另算 {k_energy_alt:.4f} N/m"]],
            ),
        },
    ]


def set_page(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.3
    ws.page_margins.bottom = 0.3
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def write_parameter_table(ws, start_row: int, start_col: int, rows: list[list[str]]) -> None:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
    title = ws.cell(start_row, start_col, rows[0][0])
    title.fill = PatternFill("solid", fgColor=HEADER_FILL)
    title.font = HEADER_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")

    for r_offset, row in enumerate(rows[1:], start=1):
        rr = start_row + r_offset
        label = ws.cell(rr, start_col, row[0])
        label.fill = PatternFill("solid", fgColor=LIGHT_FILL)
        label.font = BODY_FONT
        label.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        ws.merge_cells(start_row=rr, start_column=start_col + 1, end_row=rr, end_column=start_col + 3)
        value = ws.cell(rr, start_col + 1, row[1])
        value.font = BODY_FONT
        value.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for rr in range(start_row, start_row + len(rows)):
        ws.row_dimensions[rr].height = 18
        for cc in range(start_col, start_col + 4):
            cell = ws.cell(rr, cc)
            cell.border = BORDER
            if rr == start_row:
                cell.border = OUTLINE


def add_image(ws, image_path: Path, anchor: str, width_px: int) -> None:
    img = XLImage(str(image_path))
    with PILImage.open(image_path) as pil:
        aspect = pil.height / pil.width
    img.width = width_px
    img.height = int(width_px * aspect)
    img.anchor = anchor
    ws.add_image(img)


def write_figure_block(ws, start_row: int, spec: dict[str, object]) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=17)
    title = ws.cell(start_row, 1, spec["title"])
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    title.fill = PatternFill("solid", fgColor=WHITE)
    for col in range(1, 18):
        ws.cell(start_row, col).border = Border(top=MEDIUM, bottom=MEDIUM)
    ws.row_dimensions[start_row].height = 24

    write_parameter_table(ws, start_row + 2, 1, spec["rows"])
    add_image(ws, spec["image"], f"F{start_row + 2}", 720)

    return start_row + 19


def write_summary_sheet(wb: Workbook, data, fits) -> None:
    ws = wb.create_sheet("结果汇总", 0)
    set_page(ws)
    for idx, width in enumerate([12, 18, 18, 18, 18, 18, 18, 18], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.merge_cells("A1:H1")
    ws["A1"] = "简谐振动实验结果汇总"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    m1_g, m2_g = data["spring_masses_g"]
    m0_spring = (m1_g + m2_g) / 3.0
    k_tm, m0_tm = k_m0_from_period(fits["水平 T²-M"])
    k_inc1, m0_inc1 = k_m0_from_period(fits["斜面1 T²-M"])
    k_inc2, m0_inc2 = k_m0_from_period(fits["斜面2 T²-M"])
    rows = [
        ["焦利秤法", f"k1={fits['弹簧1 F-ΔL'].slope:.4f} N/m", f"k2={fits['弹簧2 F-ΔL'].slope:.4f} N/m", f"k1+k2={fits['弹簧1 F-ΔL'].slope + fits['弹簧2 F-ΔL'].slope:.4f} N/m"],
        ["水平 T²-M", f"K={k_tm:.4f} N/m", f"m0={m0_tm * 1000:.3f} g", "周期法"],
        ["斜面1 T²-M", f"K={k_inc1:.4f} N/m", f"m0={m0_inc1 * 1000:.3f} g", "h=1.240 cm, θ=0.822°"],
        ["斜面2 T²-M", f"K={k_inc2:.4f} N/m", f"m0={m0_inc2 * 1000:.3f} g", "h=2.510 cm, θ=1.665°"],
        ["A-T", f"Tmean={np.mean(data['amp_t_ms']):.3f} ms", f"s={np.std(data['amp_t_ms'], ddof=1):.3f} ms", "周期基本与振幅无关"],
        ["m0", f"(m1+m2)/3={m0_spring:.3f} g", "用于能量法对比", ""],
    ]
    headers = ["项目", "结果1", "结果2", "说明"]
    for c, header in enumerate(headers, start=1):
        cell = ws.cell(3, c, header)
        cell.font = HEADER_FONT
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER
    for r, row in enumerate(rows, start=4):
        for c, value in enumerate(row, start=1):
            cell = ws.cell(r, c, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = BORDER
        ws.row_dimensions[r].height = 24
    ws.print_area = "A1:H11"


def build_workbook() -> None:
    data, fits = load_data_and_results()
    specs = figure_specs(data, fits)

    wb = Workbook()
    wb.remove(wb.active)
    write_summary_sheet(wb, data, fits)

    widths = [11, 10, 13, 13, 3, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11, 11]
    for idx, spec in enumerate(specs, start=1):
        ws = wb.create_sheet(f"图{idx}")
        set_page(ws)
        for col_idx, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        for row in range(1, 25):
            ws.row_dimensions[row].height = 18
        ws.merge_cells("A1:Q1")
        ws["A1"] = "简谐振动实验图表与详细图例（参考样式）"
        ws["A1"].font = Font(name=FONT_NAME, bold=True, size=16, color=BLACK)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30
        end_row = write_figure_block(ws, 3, spec)
        ws.print_area = f"A1:Q{end_row - 1}"

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    build_workbook()
