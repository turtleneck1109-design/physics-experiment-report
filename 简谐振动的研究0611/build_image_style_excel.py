from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage


ROOT = Path(__file__).resolve().parent
FIG_DIR = ROOT / "output_figures"
OUTPUT_XLSX = ROOT / "简谐振动实验数据处理结果_图片版.xlsx"


IMAGE_SPECS = [
    ("结果汇总", "00_results_summary.png"),
    ("弹簧1拟合", "01_static_spring1_fit.png"),
    ("弹簧2拟合", "02_static_spring2_fit.png"),
    ("水平周期", "03_mass_period_fit.png"),
    ("斜面1周期", "04_incline1_mass_period_fit.png"),
    ("斜面2周期", "05_incline2_mass_period_fit.png"),
    ("振幅周期", "06_amplitude_period_fit.png"),
    ("能量关系", "07_energy_vmax2_a2_fit.png"),
]


def set_print_page(ws) -> None:
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.2
    ws.page_margins.right = 0.2
    ws.page_margins.top = 0.25
    ws.page_margins.bottom = 0.25
    ws.page_margins.header = 0.1
    ws.page_margins.footer = 0.1


def add_image_sheet(wb: Workbook, title: str, image_path: Path) -> None:
    ws = wb.create_sheet(title)
    set_print_page(ws)
    for col in range(1, 15):
        ws.column_dimensions[get_column_letter(col)].width = 12
    for row in range(1, 38):
        ws.row_dimensions[row].height = 18

    ws.merge_cells("A1:N1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    image = XLImage(str(image_path))
    with PILImage.open(image_path) as im:
        aspect = im.height / im.width

    image.width = 980
    image.height = int(image.width * aspect)
    image.anchor = "A2"
    ws.add_image(image)
    ws.print_area = "A1:N37"


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for title, filename in IMAGE_SPECS:
        add_image_sheet(wb, title, FIG_DIR / filename)
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
