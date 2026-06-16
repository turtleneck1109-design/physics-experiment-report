from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import ScatterChart, Series, Reference
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PILImage

from process_shm_data import DATA_FILE, G, linear_fit, row_values


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT / "output_figures"
OUTPUT_XLSX = ROOT / "简谐振动实验数据处理结果_打印版.xlsx"


HEADER_FILL = PatternFill("solid", fgColor="DCE8F6")
SUBHEADER_FILL = PatternFill("solid", fgColor="EDF3FB")
ALT_FILL = PatternFill("solid", fgColor="F6F8FC")
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=18)
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11)
BODY_FONT = Font(name="Microsoft YaHei", size=10)
SMALL_FONT = Font(name="Microsoft YaHei", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="AEB9C8"),
    right=Side(style="thin", color="AEB9C8"),
    top=Side(style="thin", color="AEB9C8"),
    bottom=Side(style="thin", color="AEB9C8"),
)


def set_print_layout(ws, orientation: str = "landscape", fit_width: int = 1, fit_height: int = 0) -> None:
    ws.page_setup.orientation = orientation
    ws.page_setup.fitToWidth = fit_width
    ws.page_setup.fitToHeight = fit_height
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.45
    ws.page_margins.bottom = 0.45
    ws.page_margins.header = 0.15
    ws.page_margins.footer = 0.15
    ws.sheet_view.showGridLines = False


def style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER


def write_table(ws, start_row: int, start_col: int, headers: list[str], rows: list[list[object]]) -> int:
    for offset, header in enumerate(headers):
        cell = ws.cell(start_row, start_col + offset, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for r_idx, row in enumerate(rows, start=start_row + 1):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(r_idx, c_idx, value)
            cell.font = BODY_FONT
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            if (r_idx - start_row) % 2 == 0:
                cell.fill = ALT_FILL
    return start_row + len(rows)


def section_title(ws, row: int, title: str, last_col: int = 8) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=last_col)
    cell = ws.cell(row, 1, title)
    cell.fill = SUBHEADER_FILL
    cell.font = Font(name="Microsoft YaHei", bold=True, size=12)
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def add_number_formats(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"


def load_results():
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["实验数据一页"]

    mass_g = row_values(ws, 4, 3, 7)
    force_n = mass_g / 1000.0 * G
    spring1_mass_g = float(ws.cell(row=5, column=3).value)
    spring2_mass_g = float(ws.cell(row=7, column=3).value)
    l1_m = row_values(ws, 6, 3, 7) / 1000.0
    l2_m = row_values(ws, 8, 3, 7) / 1000.0
    dl1_m = l1_m - l1_m[0]
    dl2_m = l2_m - l2_m[0]

    oscillator_mass_kg = row_values(ws, 13, 3, 7) / 1000.0
    period_s = row_values(ws, 14, 3, 7) / 1000.0

    amplitude_cm = row_values(ws, 18, 3, 6)
    amplitude_period_ms = row_values(ws, 19, 3, 6)

    amplitude_energy_cm = row_values(ws, 24, 3, 6)
    block_time_s = row_values(ws, 25, 3, 6) / 1000.0
    delta_x_m = 0.995 / 100.0
    vmax_m_s = delta_x_m / block_time_s

    incline1_mass_kg = row_values(ws, 29, 4, 7) / 1000.0
    incline1_period_s = row_values(ws, 30, 4, 7) / 1000.0
    incline2_mass_kg = row_values(ws, 31, 4, 7) / 1000.0
    incline2_period_s = row_values(ws, 32, 4, 7) / 1000.0

    fits = {
        "弹簧1 F-ΔL": linear_fit(dl1_m, force_n),
        "弹簧2 F-ΔL": linear_fit(dl2_m, force_n),
        "水平 T²-M": linear_fit(oscillator_mass_kg, period_s**2),
        "斜面1 T²-M": linear_fit(incline1_mass_kg, incline1_period_s**2),
        "斜面2 T²-M": linear_fit(incline2_mass_kg, incline2_period_s**2),
        "A-T": linear_fit(amplitude_cm, amplitude_period_ms),
        "Vmax²-A²": linear_fit((amplitude_energy_cm / 100.0) ** 2, vmax_m_s**2),
    }

    k1 = fits["弹簧1 F-ΔL"].slope
    k2 = fits["弹簧2 F-ΔL"].slope
    m0_spring_kg = (spring1_mass_g + spring2_mass_g) / 3.0 / 1000.0

    def k_m0(fit_key: str) -> tuple[float, float]:
        fit = fits[fit_key]
        return 4 * math.pi**2 / fit.slope, fit.intercept / fit.slope

    k_tm, m0_tm_kg = k_m0("水平 T²-M")
    k_inc1, m0_inc1_kg = k_m0("斜面1 T²-M")
    k_inc2, m0_inc2_kg = k_m0("斜面2 T²-M")
    k_energy_tm = fits["Vmax²-A²"].slope * (oscillator_mass_kg[0] + m0_tm_kg)
    k_energy_spring = fits["Vmax²-A²"].slope * (oscillator_mass_kg[0] + m0_spring_kg)

    data = {
        "static": (mass_g, force_n, l1_m, dl1_m, l2_m, dl2_m),
        "mass_period": (oscillator_mass_kg, period_s),
        "amplitude_period": (amplitude_cm, amplitude_period_ms),
        "energy": (amplitude_energy_cm, block_time_s, vmax_m_s),
        "inclines": (incline1_mass_kg, incline1_period_s, incline2_mass_kg, incline2_period_s),
    }
    summary_rows = [
        ["焦利秤法", f"k1 = {k1:.4f} N/m；k2 = {k2:.4f} N/m；k1+k2 = {k1 + k2:.4f} N/m", f"m0=(m1+m2)/3 = {m0_spring_kg * 1000:.3f} g"],
        ["水平 T²-M", f"K = {k_tm:.4f} N/m；m0 = {m0_tm_kg * 1000:.3f} g", f"T² = {fits['水平 T²-M'].intercept:.6f} + {fits['水平 T²-M'].slope:.6f} M"],
        ["斜面1 T²-M", f"K = {k_inc1:.4f} N/m；m0 = {m0_inc1_kg * 1000:.3f} g", "h=1.240 cm，θ=0.822°"],
        ["斜面2 T²-M", f"K = {k_inc2:.4f} N/m；m0 = {m0_inc2_kg * 1000:.3f} g", "h=2.510 cm，θ=1.665°"],
        ["A-T 关系", f"T_mean = {np.mean(amplitude_period_ms):.3f} ms；s_T = {np.std(amplitude_period_ms, ddof=1):.3f} ms", f"线性斜率 = {fits['A-T'].slope:.6f} ms/cm；周期基本与振幅无关"],
        ["Vmax²-A²", f"拟合斜率 = {fits['Vmax²-A²'].slope:.4f} s^-2；K = {k_energy_tm:.4f} N/m", f"使用水平 T²-M 的 m0；若用(m1+m2)/3，则 K={k_energy_spring:.4f} N/m"],
    ]
    return fits, data, summary_rows


def build_summary_sheet(wb: Workbook, summary_rows: list[list[str]]) -> None:
    ws = wb.active
    ws.title = "打印汇总"
    set_print_layout(ws, "landscape", 1, 1)
    ws.merge_cells("A1:H1")
    ws["A1"] = "简谐振动实验数据处理结果"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34
    ws["A2"] = "数据源：简谐振动实验数据整理_一页.xlsx；拟合方式：不加权最小二乘。"
    ws.merge_cells("A2:H2")
    ws["A2"].font = SMALL_FONT

    write_table(ws, 4, 1, ["数据处理项目", "拟合/计算结果", "备注"], summary_rows)

    section_title(ws, 12, "主要公式", 8)
    formulas = [
        ["焦利秤法", "F = k ΔL"],
        ["质量-周期法", "T² = 4π²/(k1+k2) · M + 4π²/(k1+k2) · m0"],
        ["动能-势能法", "Vmax² = (k1+k2)/(M+m0) · A²，其中 Vmax = Δx/t，Δx = 0.995 cm"],
    ]
    write_table(ws, 13, 1, ["项目", "公式"], formulas)

    section_title(ws, 18, "打印说明", 8)
    notes = [
        ["本工作簿已按页面宽度缩放，图表页建议横向打印。"],
        ["拟合参数和处理后数据可作为报告中的计算依据。"],
        ["各图表页已插入 Python 生成的 PNG 图，可直接打印或复制到报告。"],
    ]
    write_table(ws, 19, 1, ["说明"], notes)

    widths = [18, 48, 48, 12, 12, 12, 12, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(4, 25):
        ws.row_dimensions[row].height = 30
    ws.print_area = "A1:H23"


def build_fit_sheet(wb: Workbook, fits) -> None:
    ws = wb.create_sheet("拟合参数")
    set_print_layout(ws, "landscape", 1, 0)
    ws.freeze_panes = "A3"
    ws.merge_cells("A1:L1")
    ws["A1"] = "线性拟合参数汇总"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    units = {
        "弹簧1 F-ΔL": ("F=a+kΔL", "N", "N/m"),
        "弹簧2 F-ΔL": ("F=a+kΔL", "N", "N/m"),
        "水平 T²-M": ("T²=a+kM", "s²", "s²/kg"),
        "斜面1 T²-M": ("T²=a+kM", "s²", "s²/kg"),
        "斜面2 T²-M": ("T²=a+kM", "s²", "s²/kg"),
        "A-T": ("T=a+kA", "ms", "ms/cm"),
        "Vmax²-A²": ("Vmax²=a+kA²", "m²/s²", "s^-2"),
    }
    rows = []
    for name, fit in fits.items():
        equation, a_unit, k_unit = units[name]
        rows.append(
            [
                name,
                equation,
                f"{fit.intercept:.6g}",
                f"{fit.intercept_se:.6g}",
                a_unit,
                f"{fit.slope:.6g}",
                f"{fit.slope_se:.6g}",
                k_unit,
                f"{fit.ss_res:.6g}",
                f"{fit.pearson_r:.6f}",
                f"{fit.r2:.6f}",
                f"{fit.adj_r2:.6f}",
            ]
        )
    write_table(
        ws,
        3,
        1,
        ["拟合项目", "方程", "截距a", "a标准误", "a单位", "斜率k", "k标准误", "k单位", "残差平方和", "Pearson r", "R²", "调整后R²"],
        rows,
    )
    widths = [18, 18, 14, 14, 12, 14, 14, 12, 16, 12, 12, 12]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(1, 12):
        ws.row_dimensions[row].height = 24
    ws.print_area = "A1:L11"


def build_data_sheet(wb: Workbook, data) -> None:
    ws = wb.create_sheet("处理后数据")
    set_print_layout(ws, "landscape", 1, 0)
    ws.freeze_panes = "A4"
    ws.merge_cells("A1:J1")
    ws["A1"] = "实验处理后数据"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")

    mass_g, force_n, l1_m, dl1_m, l2_m, dl2_m = data["static"]
    rows = []
    for i in range(len(mass_g)):
        rows.append([i + 1, mass_g[i], force_n[i], l1_m[i] * 1000, dl1_m[i], l2_m[i] * 1000, dl2_m[i]])
    section_title(ws, 3, "一、焦利秤法", 10)
    end = write_table(ws, 4, 1, ["序号", "砝码质量/g", "F/N", "L1/mm", "ΔL1/m", "L2/mm", "ΔL2/m"], rows)

    m, t = data["mass_period"]
    rows = [[i + 1, m[i], t[i], t[i] ** 2] for i in range(len(m))]
    section_title(ws, end + 2, "二、水平状态：振子质量 M 与周期 T", 10)
    end = write_table(ws, end + 3, 1, ["序号", "M/kg", "T/s", "T²/s²"], rows)

    amp_cm, amp_t_ms = data["amplitude_period"]
    rows = [[i + 1, amp_cm[i], amp_t_ms[i]] for i in range(len(amp_cm))]
    section_title(ws, end + 2, "三、振幅 A 与周期 T", 10)
    end = write_table(ws, end + 3, 1, ["序号", "A/cm", "T/ms"], rows)

    amp_e_cm, block_t_s, vmax = data["energy"]
    rows = [[i + 1, amp_e_cm[i], (amp_e_cm[i] / 100) ** 2, block_t_s[i] * 1000, vmax[i], vmax[i] ** 2] for i in range(len(amp_e_cm))]
    section_title(ws, end + 2, "四、振幅 A 与最大速度 Vmax", 10)
    end = write_table(ws, end + 3, 1, ["序号", "A/cm", "A²/m²", "挡光t/ms", "Vmax/(m/s)", "Vmax²/(m²/s²)"], rows)

    inc1_m, inc1_t, inc2_m, inc2_t = data["inclines"]
    rows = []
    for i in range(len(inc1_m)):
        rows.append([i + 1, inc1_m[i], inc1_t[i], inc1_t[i] ** 2, inc2_m[i], inc2_t[i], inc2_t[i] ** 2])
    section_title(ws, end + 2, "五、斜面状态：振子质量 M 与周期 T", 10)
    end = write_table(ws, end + 3, 1, ["序号", "斜面1 M/kg", "斜面1 T/s", "斜面1 T²/s²", "斜面2 M/kg", "斜面2 T/s", "斜面2 T²/s²"], rows)

    widths = [10, 16, 16, 16, 16, 16, 16, 14, 14, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for row in range(1, end + 3):
        ws.row_dimensions[row].height = 22
    add_number_formats(ws)
    ws.print_area = f"A1:J{end + 1}"


def add_image_sheet(wb: Workbook, sheet_name: str, title: str, image_path: Path) -> None:
    ws = wb.create_sheet(sheet_name)
    set_print_layout(ws, "landscape", 1, 1)
    ws.merge_cells("A1:J1")
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 13
    for row in range(2, 35):
        ws.row_dimensions[row].height = 18

    image = XLImage(str(image_path))
    with PILImage.open(image_path) as pil_img:
        aspect = pil_img.height / pil_img.width
    image.width = 930
    image.height = int(image.width * aspect)
    image.anchor = "A3"
    ws.add_image(image)
    ws.print_area = "A1:J34"


def add_all_images_sheet(wb: Workbook, image_specs: list[tuple[str, str, Path]]) -> None:
    ws = wb.create_sheet("图表总览")
    set_print_layout(ws, "landscape", 1, 1)
    ws.merge_cells("A1:P1")
    ws["A1"] = "Python 作图结果总览"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 9.0
    for row in range(1, 54):
        ws.row_dimensions[row].height = 13.8

    anchors = [
        ("A3", "I3"),
        ("A15", "I15"),
        ("A27", "I27"),
        ("A39", "I39"),
    ]
    image_width = 340

    for idx, (_sheet_name, title, image_path) in enumerate(image_specs):
        row_group = idx // 2
        col_group = idx % 2
        title_cell = ws[anchors[row_group][col_group]]
        title_cell.value = title
        title_cell.font = Font(name="Microsoft YaHei", bold=True, size=10)
        title_cell.alignment = Alignment(horizontal="left", vertical="center")

        image = XLImage(str(image_path))
        with PILImage.open(image_path) as pil_img:
            aspect = pil_img.height / pil_img.width
        image.width = image_width
        image.height = int(image.width * aspect)
        title_row = 3 + row_group * 12
        image.anchor = anchors[row_group][col_group].replace(str(title_row), str(title_row + 1))
        ws.add_image(image)

    ws.print_area = "A1:P53"


def chart_data_specs(data, fits):
    mass_g, force_n, _l1_m, dl1_m, _l2_m, dl2_m = data["static"]
    m, t = data["mass_period"]
    amp_cm, amp_t_ms = data["amplitude_period"]
    amp_e_cm, block_t_s, vmax = data["energy"]
    inc1_m, inc1_t, inc2_m, inc2_t = data["inclines"]

    datasets = [
        ("弹簧1 F-ΔL", "焦利秤法：弹簧1", "ΔL / m", "F / N", dl1_m, force_n),
        ("弹簧2 F-ΔL", "焦利秤法：弹簧2", "ΔL / m", "F / N", dl2_m, force_n),
        ("水平 T²-M", "水平状态：T^2-M", "M / kg", "T^2 / s^2", m, t**2),
        ("斜面1 T²-M", "斜面1：T^2-M", "M / kg", "T^2 / s^2", inc1_m, inc1_t**2),
        ("斜面2 T²-M", "斜面2：T^2-M", "M / kg", "T^2 / s^2", inc2_m, inc2_t**2),
        ("A-T", "振幅 A 与周期 T", "A / cm", "T / ms", amp_cm, amp_t_ms),
        ("Vmax²-A²", "Vmax^2-A^2", "A^2 / m^2", "Vmax^2 / (m^2/s^2)", (amp_e_cm / 100.0) ** 2, vmax**2),
    ]

    # Keep the summary chart count at eight, but avoid rasterized tables: this
    # eighth native chart compares the three K values obtained from T^2-M data.
    k_names = np.array([1, 2, 3], dtype=float)
    k_values = np.array(
        [
            4 * math.pi**2 / fits["水平 T²-M"].slope,
            4 * math.pi**2 / fits["斜面1 T²-M"].slope,
            4 * math.pi**2 / fits["斜面2 T²-M"].slope,
        ],
        dtype=float,
    )
    datasets.append(("K对比", "周期法 K 值对比", "序号（1水平，2斜面1，3斜面2）", "K / (N/m)", k_names, k_values))
    return datasets


def build_chart_data_sheet(wb: Workbook, data, fits) -> dict[str, dict[str, object]]:
    ws = wb.create_sheet("图表数据")
    ws.sheet_state = "hidden"
    mapping: dict[str, dict[str, object]] = {}
    row = 1
    for fit_key, title, x_title, y_title, x_values, y_values in chart_data_specs(data, fits):
        fit = fits.get(fit_key)
        ws.cell(row, 1, title)
        ws.cell(row + 1, 1, "x")
        ws.cell(row + 1, 2, "实验数据")
        ws.cell(row + 1, 3, "拟合线")
        for idx, (x, y) in enumerate(zip(x_values, y_values), start=row + 2):
            ws.cell(idx, 1, float(x))
            ws.cell(idx, 2, float(y))
            ws.cell(idx, 3, float(fit.intercept + fit.slope * x) if fit else None)
        n = len(x_values)
        mapping[title] = {
            "fit_key": fit_key,
            "row": row + 2,
            "n": n,
            "x_title": x_title,
            "y_title": y_title,
        }
        row += n + 4
    return mapping


def add_scatter_chart(ws, data_ws, anchor: str, title: str, info: dict[str, object]) -> None:
    start_row = int(info["row"])
    n = int(info["n"])
    end_row = start_row + n - 1
    chart = ScatterChart()
    chart.title = title
    chart.style = 13
    chart.x_axis.title = str(info["x_title"])
    chart.y_axis.title = str(info["y_title"])
    chart.legend.position = "b"
    chart.width = 10.0
    chart.height = 4.2

    x_ref = Reference(data_ws, min_col=1, min_row=start_row, max_row=end_row)
    y_ref = Reference(data_ws, min_col=2, min_row=start_row, max_row=end_row)
    data_series = Series(y_ref, x_ref, title="实验数据")
    data_series.marker.symbol = "circle"
    data_series.marker.size = 5
    data_series.graphicalProperties.line.noFill = True
    data_series.graphicalProperties.solidFill = "2266AA"
    data_series.marker.graphicalProperties.solidFill = "2266AA"
    data_series.marker.graphicalProperties.line.solidFill = "2266AA"
    chart.series.append(data_series)

    if str(info["fit_key"]) != "K对比":
        fit_ref = Reference(data_ws, min_col=3, min_row=start_row, max_row=end_row)
        fit_series = Series(fit_ref, x_ref, title="线性拟合")
        fit_series.marker.symbol = "none"
        fit_series.graphicalProperties.line.solidFill = "C7433E"
        fit_series.graphicalProperties.line.width = 22000
        chart.series.append(fit_series)

    ws.add_chart(chart, anchor)


def add_native_charts_sheet(wb: Workbook, chart_mapping: dict[str, dict[str, object]]) -> None:
    ws = wb.create_sheet("图表总览")
    set_print_layout(ws, "landscape", 1, 1)
    ws.merge_cells("A1:P1")
    ws["A1"] = "Excel 原生图表总览"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    for col in range(1, 17):
        ws.column_dimensions[get_column_letter(col)].width = 8.5
    for row in range(1, 57):
        ws.row_dimensions[row].height = 14.5

    data_ws = wb["图表数据"]
    placements = [
        ("A2", "I2"),
        ("A16", "I16"),
        ("A30", "I30"),
        ("A44", "I44"),
    ]
    titles = [
        "焦利秤法：弹簧1",
        "焦利秤法：弹簧2",
        "水平状态：T^2-M",
        "斜面1：T^2-M",
        "斜面2：T^2-M",
        "振幅 A 与周期 T",
        "Vmax^2-A^2",
        "周期法 K 值对比",
    ]
    for idx, title in enumerate(titles):
        add_scatter_chart(ws, data_ws, placements[idx // 2][idx % 2], title, chart_mapping[title])
    ws.print_area = "A1:P57"


def main() -> None:
    fits, data, summary_rows = load_results()
    wb = Workbook()
    build_summary_sheet(wb, summary_rows)
    build_fit_sheet(wb, fits)
    build_data_sheet(wb, data)
    chart_mapping = build_chart_data_sheet(wb, data, fits)

    image_specs = [
        ("图0汇总", "结果汇总图", OUT_DIR / "00_results_summary.png"),
        ("图1弹簧1", "焦利秤法：弹簧1", OUT_DIR / "01_static_spring1_fit.png"),
        ("图2弹簧2", "焦利秤法：弹簧2", OUT_DIR / "02_static_spring2_fit.png"),
        ("图3水平周期", "水平状态：T²-M 拟合", OUT_DIR / "03_mass_period_fit.png"),
        ("图4斜面1", "斜面1：T²-M 拟合", OUT_DIR / "04_incline1_mass_period_fit.png"),
        ("图5斜面2", "斜面2：T²-M 拟合", OUT_DIR / "05_incline2_mass_period_fit.png"),
        ("图6振幅周期", "振幅 A 与周期 T", OUT_DIR / "06_amplitude_period_fit.png"),
        ("图7能量关系", "Vmax²-A² 拟合", OUT_DIR / "07_energy_vmax2_a2_fit.png"),
    ]
    add_native_charts_sheet(wb, chart_mapping)

    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
