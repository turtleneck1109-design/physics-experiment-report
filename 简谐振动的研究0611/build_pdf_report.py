from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import openpyxl
from PIL import Image as PILImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    Image,
    PageBreak,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from process_shm_data import G, DATA_FILE, linear_fit, row_values


ROOT = Path(__file__).resolve().parent
FIG_DIR = ROOT / "output_figures"
OUT_DIR = ROOT / "output" / "pdf"
PDF_PATH = OUT_DIR / "简谐振动实验数据处理结果汇报.pdf"


def setup_styles():
    pdfmetrics.registerFont(UnicodeCIDFont("STSong-Light"))
    styles = getSampleStyleSheet()
    base = "STSong-Light"
    return {
        "title": ParagraphStyle(
            "TitleCN",
            parent=styles["Title"],
            fontName=base,
            fontSize=22,
            leading=28,
            alignment=1,
            spaceAfter=8,
        ),
        "subtitle": ParagraphStyle(
            "SubtitleCN",
            parent=styles["Normal"],
            fontName=base,
            fontSize=10,
            leading=14,
            alignment=1,
            textColor=colors.HexColor("#444444"),
            spaceAfter=12,
        ),
        "h1": ParagraphStyle(
            "HeadingCN",
            parent=styles["Heading1"],
            fontName=base,
            fontSize=14,
            leading=18,
            spaceBefore=8,
            spaceAfter=6,
        ),
        "body": ParagraphStyle(
            "BodyCN",
            parent=styles["BodyText"],
            fontName=base,
            fontSize=9.5,
            leading=14,
            spaceAfter=4,
        ),
        "small": ParagraphStyle(
            "SmallCN",
            parent=styles["BodyText"],
            fontName=base,
            fontSize=8.3,
            leading=11,
        ),
        "caption": ParagraphStyle(
            "CaptionCN",
            parent=styles["BodyText"],
            fontName=base,
            fontSize=9,
            leading=12,
            alignment=1,
            textColor=colors.HexColor("#333333"),
            spaceAfter=5,
        ),
    }


def fit_results():
    wb = openpyxl.load_workbook(DATA_FILE, data_only=True)
    ws = wb["实验数据一页"]

    mass_g = row_values(ws, 4, 3, 7)
    force_n = mass_g / 1000.0 * G
    spring1_mass_g = float(ws.cell(row=5, column=3).value)
    spring2_mass_g = float(ws.cell(row=7, column=3).value)
    l1_m = row_values(ws, 6, 3, 7) / 1000.0
    l2_m = row_values(ws, 8, 3, 7) / 1000.0
    dl1_m = l1_m - l1_m[0]
    dl2_m = l2_m - l2_m[0]

    m_kg = row_values(ws, 13, 3, 7) / 1000.0
    t_s = row_values(ws, 14, 3, 7) / 1000.0
    amp_cm = row_values(ws, 18, 3, 6)
    amp_t_ms = row_values(ws, 19, 3, 6)
    amp_energy_cm = row_values(ws, 24, 3, 6)
    block_t_s = row_values(ws, 25, 3, 6) / 1000.0
    vmax = (0.995 / 100.0) / block_t_s
    inc1_m = row_values(ws, 29, 4, 7) / 1000.0
    inc1_t = row_values(ws, 30, 4, 7) / 1000.0
    inc2_m = row_values(ws, 31, 4, 7) / 1000.0
    inc2_t = row_values(ws, 32, 4, 7) / 1000.0

    fits = {
        "spring1": linear_fit(dl1_m, force_n),
        "spring2": linear_fit(dl2_m, force_n),
        "mass_period": linear_fit(m_kg, t_s**2),
        "inc1": linear_fit(inc1_m, inc1_t**2),
        "inc2": linear_fit(inc2_m, inc2_t**2),
        "amp_period": linear_fit(amp_cm, amp_t_ms),
        "energy": linear_fit((amp_energy_cm / 100.0) ** 2, vmax**2),
    }

    k1 = fits["spring1"].slope
    k2 = fits["spring2"].slope
    m0_spring_g = (spring1_mass_g + spring2_mass_g) / 3.0

    def k_m0(key: str) -> tuple[float, float]:
        fit = fits[key]
        return 4 * math.pi**2 / fit.slope, fit.intercept / fit.slope * 1000

    k_tm, m0_tm_g = k_m0("mass_period")
    k_inc1, m0_inc1_g = k_m0("inc1")
    k_inc2, m0_inc2_g = k_m0("inc2")
    k_energy = fits["energy"].slope * (m_kg[0] + m0_tm_g / 1000)

    summary = [
        ["焦利秤法", f"k1={k1:.4f} N/m, k2={k2:.4f} N/m", f"k1+k2={k1+k2:.4f} N/m; m0={m0_spring_g:.3f} g"],
        ["水平 T^2-M", f"K={k_tm:.4f} N/m", f"m0={m0_tm_g:.3f} g"],
        ["斜面1 T^2-M", f"K={k_inc1:.4f} N/m", f"m0={m0_inc1_g:.3f} g; theta=0.822 deg"],
        ["斜面2 T^2-M", f"K={k_inc2:.4f} N/m", f"m0={m0_inc2_g:.3f} g; theta=1.665 deg"],
        ["倾角影响", "周期与斜面倾角无明显关系", "倾角主要改变平衡位置"],
        ["A-T 关系", f"Tmean={np.mean(amp_t_ms):.3f} ms", f"s={np.std(amp_t_ms, ddof=1):.3f} ms; 周期基本与振幅无关"],
        ["Vmax^2-A^2", f"斜率={fits['energy'].slope:.4f} s^-2", f"K={k_energy:.4f} N/m"],
    ]

    return summary


def make_table(rows, col_widths, header=None, font_size=8.5):
    data = ([header] if header else []) + rows
    table = Table(data, colWidths=col_widths, repeatRows=1 if header else 0)
    style = [
        ("FONTNAME", (0, 0), (-1, -1), "STSong-Light"),
        ("FONTSIZE", (0, 0), (-1, -1), font_size),
        ("LEADING", (0, 0), (-1, -1), font_size + 2),
        ("GRID", (0, 0), (-1, -1), 0.45, colors.HexColor("#9CA3AF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
    if header:
        style.extend(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8EEF7")),
                ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ]
        )
    for row_idx in range(1 if header else 0, len(data)):
        if row_idx % 2 == 0:
            style.append(("BACKGROUND", (0, row_idx), (-1, row_idx), colors.HexColor("#F7F9FC")))
    table.setStyle(TableStyle(style))
    return table


def scaled_image(path: Path, max_width: float, max_height: float) -> Image:
    with PILImage.open(path) as im:
        width, height = im.size
    scale = min(max_width / width, max_height / height)
    return Image(str(path), width=width * scale, height=height * scale)


def footer(canvas, doc):
    canvas.saveState()
    canvas.setFont("STSong-Light", 8)
    canvas.setFillColor(colors.HexColor("#666666"))
    canvas.drawCentredString(A4[0] / 2, 10 * mm, f"第 {doc.page} 页")
    canvas.restoreState()


def build_pdf() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    styles = setup_styles()
    doc = SimpleDocTemplate(
        str(PDF_PATH),
        pagesize=A4,
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=14 * mm,
        bottomMargin=16 * mm,
        title="简谐振动实验数据处理结果汇报",
    )

    story = [
        Paragraph("简谐振动实验数据处理结果汇报", styles["title"]),
        Paragraph("数据源：简谐振动实验数据整理_一页.xlsx　　作图与拟合：Python", styles["subtitle"]),
        Paragraph("一、处理方法", styles["h1"]),
        make_table(
            [
                ["焦利秤法", "F = k Delta L，用 F-Delta L 线性拟合求弹簧劲度系数。"],
                ["质量-周期法", "T^2 = 4pi^2/(k1+k2) M + 4pi^2/(k1+k2) m0，拟合 T^2-M 求 K 与 m0。"],
                ["能量关系", "Vmax^2 = (k1+k2)/(M+m0) A^2，其中 Vmax = Delta x / t，Delta x = 0.995 cm。"],
            ],
            [35 * mm, 135 * mm],
            header=["项目", "处理公式/说明"],
            font_size=8.2,
        ),
        Spacer(1, 5 * mm),
        Paragraph("二、主要结果", styles["h1"]),
        make_table(
            fit_results(),
            [30 * mm, 58 * mm, 82 * mm],
            header=["数据处理项目", "拟合/计算结果", "说明"],
            font_size=8.2,
        ),
        Spacer(1, 5 * mm),
        Paragraph("三、结论", styles["h1"]),
        Paragraph("1. 焦利秤法得到总劲度系数 k1+k2=4.7461 N/m；周期法得到水平状态 K=4.8624 N/m，斜面1 K=4.9087 N/m，斜面2 K=4.9035 N/m。两种方法得到的总劲度系数整体接近。", styles["body"]),
        Paragraph("2. 斜面倾角改变时，同质量下周期和拟合得到的 K 值变化很小，可认为周期与斜面倾角无明显关系。", styles["body"]),
        Paragraph("3. A-T 关系的线性相关性很弱，说明在本实验范围内周期基本与振幅无关。", styles["body"]),
        Paragraph("4. Vmax^2-A^2 呈明显线性关系，符合简谐运动动能与势能转化关系。", styles["body"]),
        PageBreak(),
    ]

    image_groups = [
        ("四、焦利秤法拟合图", ["01_static_spring1_fit.png", "02_static_spring2_fit.png"]),
        ("五、质量与周期关系拟合图", ["03_mass_period_fit.png", "04_incline1_mass_period_fit.png"]),
        ("六、斜面与振幅相关拟合图", ["05_incline2_mass_period_fit.png", "06_amplitude_period_fit.png"]),
        ("七、能量关系拟合图", ["07_energy_vmax2_a2_fit.png"]),
    ]

    for group_idx, (heading, filenames) in enumerate(image_groups):
        story.append(Paragraph(heading, styles["h1"]))
        for filename in filenames:
            path = FIG_DIR / filename
            story.append(scaled_image(path, doc.width, 88 * mm))
            story.append(Spacer(1, 3 * mm))
        if group_idx != len(image_groups) - 1:
            story.append(PageBreak())

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    print(PDF_PATH)


if __name__ == "__main__":
    build_pdf()
