"""生成实验数据处理结果。

这个脚本负责三件事：
1. 从原始 Excel 读取实验数据；
2. 完成平均值、最小二乘拟合、杨氏模量和不确定度计算；
3. 输出 matplotlib 拟合图和整理后的 Excel 表格。
"""

from __future__ import annotations

import math
from pathlib import Path
from copy import copy
from statistics import mean, stdev

import matplotlib

# 使用无界面后端，只把图保存成文件，不弹出绘图窗口。
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(__file__).resolve().parent
INPUT_XLSX = ROOT / "数据处理.xlsx"
OUTPUT_DIR = ROOT / "outputs"
PLOT_PATH = OUTPUT_DIR / "m_b_linear_fit_matplotlib.png"
OUTPUT_XLSX = ROOT / "数据处理_最终表格.xlsx"

G = 9.8
RULER_U_M = 0.0005  # 米尺仪器不确定度，单位 m
MICROMETER_U_MM = 0.005  # 螺旋测微器仪器不确定度，单位 mm


def f(value) -> float:
    """把 Excel 中可能以文本存储的数字转成 float。"""
    return float(value)


def direct_uncertainty(values: list[float], instrument_u: float) -> dict[str, float]:
    """计算直接测量量的不确定度。

    PPT 要求：
    ΔA = 1.05σ
    ΔB = 仪器不确定度
    u = sqrt(ΔA^2 + ΔB^2)
    """
    avg = mean(values)
    sigma = stdev(values) if len(values) > 1 else 0.0
    delta_a = 1.05 * sigma
    delta_b = instrument_u
    u = math.sqrt(delta_a**2 + delta_b**2)
    return {
        "mean": avg,
        "sigma": sigma,
        "delta_a": delta_a,
        "delta_b": delta_b,
        "u": u,
    }


def linear_fit(x: list[float], y: list[float]) -> dict[str, float | list[float]]:
    """用最小二乘法拟合 y = intercept + slope * x。

    本实验中：
    x = 标尺读数 b/cm
    y = 砝码质量 m/kg
    slope 就是后续计算 E 所需的 k。
    """
    n = len(x)
    x_bar = mean(x)
    y_bar = mean(y)
    sxx = sum((xi - x_bar) ** 2 for xi in x)
    syy = sum((yi - y_bar) ** 2 for yi in y)
    sxy = sum((xi - x_bar) * (yi - y_bar) for xi, yi in zip(x, y))
    slope = sxy / sxx
    intercept = y_bar - slope * x_bar
    fitted = [intercept + slope * xi for xi in x]
    residuals = [yi - fi for yi, fi in zip(y, fitted)]
    ssr = sum(ri**2 for ri in residuals)
    residual_variance = ssr / (n - 2)  # 线性拟合有斜率和截距两个参数，自由度为 n-2
    slope_u = math.sqrt(residual_variance / sxx)
    intercept_u = math.sqrt(residual_variance * (1 / n + x_bar**2 / sxx))
    pearson_r = sxy / math.sqrt(sxx * syy)
    r2 = pearson_r**2
    adj_r2 = 1 - (1 - r2) * (n - 1) / (n - 2)
    return {
        "n": n,
        "x_bar": x_bar,
        "y_bar": y_bar,
        "sxx": sxx,
        "syy": syy,
        "sxy": sxy,
        "slope": slope,
        "intercept": intercept,
        "slope_u": slope_u,
        "intercept_u": intercept_u,
        "ssr": ssr,
        "residual_variance": residual_variance,
        "pearson_r": pearson_r,
        "r2": r2,
        "adj_r2": adj_r2,
        "fitted": fitted,
        "residuals": residuals,
    }


def load_data() -> dict:
    """从 Sheet1 读取原始数据，并计算每个砝码对应的平均 b。"""
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=False)
    ws = wb["Sheet1"]
    data = {
        "L_values": [f(ws.cell(row=3, column=c).value) for c in range(2, 8)],
        "l_values": [f(ws.cell(row=4, column=c).value) for c in range(2, 8)],
        "D_values": [f(ws.cell(row=5, column=c).value) for c in range(2, 8)],
        "d_values": [f(ws.cell(row=6, column=c).value) for c in range(2, 8)],
        "mass": [f(ws.cell(row=8, column=c).value) for c in range(2, 10)],
        "add": [f(ws.cell(row=9, column=c).value) for c in range(2, 10)],
        "remove": [f(ws.cell(row=10, column=c).value) for c in range(2, 10)],
    }
    # PPT 表格要求：b 取加砝码读数和减砝码读数的平均值。
    data["b"] = [(a + r) / 2 for a, r in zip(data["add"], data["remove"])]
    return data


def analyze(data: dict) -> dict:
    """完成全部物理量计算，并把结果集中放进一个 dict。"""
    L = direct_uncertainty(data["L_values"], RULER_U_M)
    small_l = direct_uncertainty(data["l_values"], RULER_U_M)
    D = direct_uncertainty(data["D_values"], RULER_U_M)
    d = direct_uncertainty(data["d_values"], MICROMETER_U_MM)
    fit = linear_fit(data["b"], data["mass"])

    # 拟合时横坐标 b 用 cm，计算 E 时要换成 SI 单位 m。
    k_kg_per_m = fit["slope"] * 100
    uk_kg_per_m = fit["slope_u"] * 100

    # 钢丝直径原始单位是 mm，代入 E 公式时必须换成 m。
    d_m = d["mean"] / 1000
    ud_m = d["u"] / 1000

    # PPT 中的公式：E = 8gDLk / (πd²l)。
    E = (8 * G / math.pi) * D["mean"] * L["mean"] * k_kg_per_m / (d_m**2 * small_l["mean"])

    # 因为 E ∝ D * L * k * d^-2 * l^-1，所以相对不确定度按幂函数规则合成。
    terms = {
        "uD/D": D["u"] / D["mean"],
        "uL/L": L["u"] / L["mean"],
        "uk/k": uk_kg_per_m / k_kg_per_m,
        "2ud/d": 2 * ud_m / d_m,
        "ul/l": small_l["u"] / small_l["mean"],
    }
    rel_u_E = math.sqrt(sum(v**2 for v in terms.values()))
    u_E = E * rel_u_E
    return {
        "L": L,
        "l": small_l,
        "D": D,
        "d": d,
        "fit": fit,
        "k_kg_per_m": k_kg_per_m,
        "uk_kg_per_m": uk_kg_per_m,
        "E": E,
        "u_E": u_E,
        "rel_u_E": rel_u_E,
        "terms": terms,
    }


def make_plot(data: dict, result: dict) -> None:
    """用 matplotlib 生成散点图、拟合直线和拟合参数表。"""
    OUTPUT_DIR.mkdir(exist_ok=True)
    plt.rcParams["font.sans-serif"] = ["Microsoft YaHei", "SimHei", "Arial Unicode MS"]
    plt.rcParams["axes.unicode_minus"] = False

    b = data["b"]
    mass = data["mass"]
    fit = result["fit"]
    slope = fit["slope"]
    intercept = fit["intercept"]

    x_min, x_max = -0.5, 9.0
    x_line = [x_min, x_max]
    y_line = [intercept + slope * x for x in x_line]

    fig, ax = plt.subplots(figsize=(10.5, 7.0), dpi=180)
    ax.scatter(b, mass, marker="s", s=34, color="#333333", label="砝码质量", zorder=3)
    ax.plot(x_line, y_line, color="#bf5146", linewidth=1.8, label="砝码质量-标尺读数的线性拟合", zorder=2)
    ax.set_title("砝码质量-标尺读数关系图", fontsize=16, fontweight="bold", pad=12)
    ax.set_xlabel("标尺读数 b/cm", fontsize=12)
    ax.set_ylabel("砝码质量 m/kg", fontsize=12)
    ax.set_xlim(x_min, x_max)
    ax.set_ylim(-0.2, 4.0)
    ax.set_xticks([0, 2, 4, 6, 8])
    ax.set_yticks([0, 0.5, 1, 1.5, 2, 2.5, 3, 3.5, 4])
    ax.grid(True, color="#dddddd", linewidth=0.6, alpha=0.8, zorder=0)
    ax.legend(loc="upper left", frameon=True, edgecolor="#999999", fontsize=10)

    table_rows = [
        ["方程", "m = a + k b"],
        ["绘图", "砝码质量"],
        ["权重", "不加权"],
        ["截距", f"{intercept:.5f} ± {fit['intercept_u']:.5f} kg"],
        ["斜率", f"{slope:.5f} ± {fit['slope_u']:.5f} kg/cm"],
        ["残差平方和", f"{fit['ssr']:.6g}"],
        ["Pearson's r", f"{fit['pearson_r']:.6f}"],
        ["R²(COD)", f"{fit['r2']:.6f}"],
        ["调整后R²", f"{fit['adj_r2']:.6f}"],
    ]
    stats_table = ax.table(
        cellText=table_rows,
        cellLoc="left",
        colWidths=[0.33, 0.67],
        bbox=[0.56, 0.08, 0.40, 0.34],
    )
    stats_table.set_zorder(5)
    stats_table.auto_set_font_size(False)
    stats_table.set_fontsize(9)
    for cell in stats_table.get_celld().values():
        cell.set_edgecolor("#8a8a8a")
        cell.set_linewidth(0.5)
        cell.set_facecolor("#f7f7f7")
        cell.set_alpha(1.0)

    fig.tight_layout()
    fig.savefig(PLOT_PATH, bbox_inches="tight")
    plt.close(fig)


def style_title(ws, title: str, last_col: str) -> None:
    """设置 Excel 工作表顶部标题样式。"""
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=16, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    ws.sheet_view.showGridLines = False


def style_block(ws, cell_range: str) -> None:
    """给 Excel 表格区域添加边框和居中换行。"""
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row in ws[cell_range]:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def header_row(ws, row: int, first_col: int, last_col: int) -> None:
    """设置 Excel 表头行样式。"""
    fill = PatternFill("solid", fgColor="1F77B4")
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def add_table(ws, ref: str, name: str) -> None:
    """把普通单元格区域转换成 Excel 表格对象，便于筛选和阅读。"""
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def write_workbook(data: dict, result: dict) -> None:
    """把计算结果、原始数据、拟合过程和拟合图写入最终 Excel 文件。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "结果汇总"
    style_title(ws, "拉伸法测金属丝杨氏模量数据处理结果", "H")

    summary = [
        ["项目", "结果", "说明"],
        ["拟合方程", f"m = {result['fit']['intercept']:.5f} + {result['fit']['slope']:.5f} b", "b 单位 cm，m 单位 kg"],
        ["斜率 k", f"{result['fit']['slope']:.5f} ± {result['fit']['slope_u']:.5f} kg/cm", "最小二乘线性拟合"],
        ["斜率 k(SI)", f"{result['k_kg_per_m']:.4f} ± {result['uk_kg_per_m']:.4f} kg/m", "计算 E 时使用"],
        ["Pearson's r", f"{result['fit']['pearson_r']:.6f}", "相关系数"],
        ["R²(COD)", f"{result['fit']['r2']:.6f}", "决定系数"],
        ["E", f"{result['E']:.4e} Pa", "E = 8gDLk/(πd²l)"],
        ["uE", f"{result['u_E']:.4e} Pa", "不确定度传递"],
        ["相对不确定度", f"{result['rel_u_E'] * 100:.2f}%", "uE/E"],
        ["最终结果", f"({result['E'] / 1e11:.2f} ± {result['u_E'] / 1e11:.2f}) × 10^11 Pa", "按实验报告格式"],
    ]
    for r, row in enumerate(summary, start=3):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    header_row(ws, 3, 1, 3)
    style_block(ws, "A3:C12")
    add_table(ws, "A3:C12", "SummaryTable")

    ws["A15"] = "直接测量量及合成不确定度"
    ws["A15"].font = Font(name="Microsoft YaHei", bold=True, color="1F4E79", size=13)
    direct_rows = [
        ["量", "单位", "平均值", "σ", "ΔA=1.05σ", "ΔB", "合成 u", "表示结果"],
        ["钢丝长度 L", "m", result["L"]["mean"], result["L"]["sigma"], result["L"]["delta_a"], result["L"]["delta_b"], result["L"]["u"], f"{result['L']['mean']:.4f} ± {result['L']['u']:.4f} m"],
        ["平面镜到钢丝距离 l", "m", result["l"]["mean"], result["l"]["sigma"], result["l"]["delta_a"], result["l"]["delta_b"], result["l"]["u"], f"{result['l']['mean']:.4f} ± {result['l']['u']:.4f} m"],
        ["平面镜到望远镜直尺距离 D", "m", result["D"]["mean"], result["D"]["sigma"], result["D"]["delta_a"], result["D"]["delta_b"], result["D"]["u"], f"{result['D']['mean']:.4f} ± {result['D']['u']:.4f} m"],
        ["钢丝直径 d", "mm", result["d"]["mean"], result["d"]["sigma"], result["d"]["delta_a"], result["d"]["delta_b"], result["d"]["u"], f"{result['d']['mean']:.3f} ± {result['d']['u']:.3f} mm"],
    ]
    for r, row in enumerate(direct_rows, start=16):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    header_row(ws, 16, 1, 8)
    style_block(ws, "A16:H20")
    add_table(ws, "A16:H20", "DirectUncertainty")

    ws["A23"] = "不确定度传递"
    ws["A23"].font = Font(name="Microsoft YaHei", bold=True, color="1F4E79", size=13)
    prop_rows = [["分量", "相对量", "平方贡献"]]
    for label, value in result["terms"].items():
        prop_rows.append([label, value, value**2])
    prop_rows.extend(
        [
            ["合成 uE/E", result["rel_u_E"], result["rel_u_E"] ** 2],
            ["uE/Pa", result["u_E"], ""],
        ]
    )
    for r, row in enumerate(prop_rows, start=24):
        for c, value in enumerate(row, start=1):
            ws.cell(r, c, value)
    header_row(ws, 24, 1, 3)
    style_block(ws, f"A24:C{23 + len(prop_rows)}")
    add_table(ws, f"A24:C{23 + len(prop_rows)}", "PropagationTable")

    for col, width in {"A": 25, "B": 24, "C": 30, "D": 14, "E": 14, "F": 14, "G": 14, "H": 24}.items():
        ws.column_dimensions[col].width = width

    raw_ws = wb.create_sheet("原始与平均")
    style_title(raw_ws, "原始数据与平均 b", "I")
    rows = [
        ["序号", "砝码质量 m/kg", "加砝码读数/cm", "减砝码读数/cm", "平均 b/cm"],
    ]
    for i, (m, add, remove, b) in enumerate(zip(data["mass"], data["add"], data["remove"], data["b"]), start=1):
        rows.append([i, m, add, remove, b])
    for r, row in enumerate(rows, start=3):
        for c, value in enumerate(row, start=1):
            raw_ws.cell(r, c, value)
    header_row(raw_ws, 3, 1, 5)
    style_block(raw_ws, "A3:E11")
    add_table(raw_ws, "A3:E11", "AverageBTable")

    raw_ws["A14"] = "直接测量原始读数"
    raw_ws["A14"].font = Font(name="Microsoft YaHei", bold=True, color="1F4E79", size=13)
    direct_raw = [
        ["测量次数", "第1次", "第2次", "第3次", "第4次", "第5次", "第6次"],
        ["钢丝长度 L/m", *data["L_values"]],
        ["平面镜到钢丝距离 l/m", *data["l_values"]],
        ["平面镜到望远镜直尺距离 D/m", *data["D_values"]],
        ["钢丝直径 d/mm", *data["d_values"]],
    ]
    for r, row in enumerate(direct_raw, start=15):
        for c, value in enumerate(row, start=1):
            raw_ws.cell(r, c, value)
    header_row(raw_ws, 15, 1, 7)
    style_block(raw_ws, "A15:G19")
    add_table(raw_ws, "A15:G19", "DirectRawTable")
    for col, width in {"A": 28, "B": 16, "C": 16, "D": 16, "E": 16, "F": 16, "G": 16}.items():
        raw_ws.column_dimensions[col].width = width

    fit_ws = wb.create_sheet("最小二乘拟合")
    style_title(fit_ws, "最小二乘线性拟合计算过程", "M")
    headers = [
        "序号",
        "b/cm",
        "m/kg",
        "拟合 m/kg",
        "b-b̄",
        "m-m̄",
        "(b-b̄)^2",
        "(m-m̄)^2",
        "(b-b̄)(m-m̄)",
        "残差",
        "残差^2",
    ]
    for c, header in enumerate(headers, start=1):
        fit_ws.cell(3, c, header)
    header_row(fit_ws, 3, 1, len(headers))
    fit = result["fit"]
    for idx, (b, m, fitted, residual) in enumerate(
        zip(data["b"], data["mass"], fit["fitted"], fit["residuals"]), start=1
    ):
        row = 3 + idx
        db = b - fit["x_bar"]
        dm = m - fit["y_bar"]
        values = [idx, b, m, fitted, db, dm, db**2, dm**2, db * dm, residual, residual**2]
        for c, value in enumerate(values, start=1):
            fit_ws.cell(row, c, value)
    style_block(fit_ws, "A3:K11")
    add_table(fit_ws, "A3:K11", "LeastSquaresAux")

    fit_summary = [
        ["项目", "数值", "说明"],
        ["n", fit["n"], "数据点数"],
        ["b̄/cm", fit["x_bar"], "横坐标平均值"],
        ["m̄/kg", fit["y_bar"], "纵坐标平均值"],
        ["Sxx", fit["sxx"], "Σ(bi-b̄)^2"],
        ["Syy", fit["syy"], "Σ(mi-m̄)^2"],
        ["Sxy", fit["sxy"], "Σ(bi-b̄)(mi-m̄)"],
        ["残差平方和", fit["ssr"], "Σ[mi-(a+kbi)]²"],
        ["残差方差", fit["residual_variance"], "RSS/(n-2)"],
        ["截距 a/kg", fit["intercept"], "m=a+kb"],
        ["u(a)/kg", fit["intercept_u"], "截距标准不确定度"],
        ["斜率 k/(kg/cm)", fit["slope"], "Sxy/Sxx"],
        ["u(k)/(kg/cm)", fit["slope_u"], "斜率标准不确定度"],
        ["k/(kg/m)", result["k_kg_per_m"], "SI 单位"],
        ["u(k)/(kg/m)", result["uk_kg_per_m"], "SI 单位"],
        ["Pearson's r", fit["pearson_r"], "相关系数"],
        ["R²(COD)", fit["r2"], "决定系数"],
        ["调整后 R²", fit["adj_r2"], "调整后决定系数"],
    ]
    for r, row in enumerate(fit_summary, start=14):
        for c, value in enumerate(row, start=1):
            fit_ws.cell(r, c, value)
    header_row(fit_ws, 14, 1, 3)
    style_block(fit_ws, f"A14:C{13 + len(fit_summary)}")
    add_table(fit_ws, f"A14:C{13 + len(fit_summary)}", "LeastSquaresSummary")
    for col, width in {"A": 18, "B": 18, "C": 24, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 20, "J": 16, "K": 16}.items():
        fit_ws.column_dimensions[col].width = width

    graph_ws = wb.create_sheet("拟合图")
    graph_ws.sheet_view.showGridLines = False
    graph_ws["A1"] = "Python matplotlib 生成的最小二乘线性拟合图"
    graph_ws["A1"].font = Font(name="Microsoft YaHei", bold=True, size=16, color="1F4E79")
    img = XLImage(str(PLOT_PATH))
    img.width = 1000
    img.height = 700
    graph_ws.add_image(img, "A3")
    graph_ws.column_dimensions["A"].width = 140

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                new_font = copy(cell.font)
                new_font.name = "Microsoft YaHei"
                cell.font = new_font
        sheet.sheet_view.selection[0].sqref = "A1"
        sheet.sheet_view.selection[0].activeCell = "A1"

    for sheet in ["结果汇总", "原始与平均", "最小二乘拟合"]:
        wb[sheet].freeze_panes = "A3"

    wb.active = 0
    wb.save(OUTPUT_XLSX)


def main() -> None:
    data = load_data()
    result = analyze(data)
    make_plot(data, result)
    write_workbook(data, result)
    print(f"workbook={OUTPUT_XLSX}")
    print(f"plot={PLOT_PATH}")
    print(f"final=({result['E'] / 1e11:.2f} ± {result['u_E'] / 1e11:.2f}) × 10^11 Pa")


if __name__ == "__main__":
    main()
